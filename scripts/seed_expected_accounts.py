"""Seed `expected_account` rows from XLSX + Credit Karma list.

Two sources:

1. **XLSX `Account Summary` sheet** (`/Users/travis/Downloads/Savings & Retirement
   Plan.xlsx`): every distinct non-aggregate account label becomes one
   `ExpectedAccount` with ``source='xlsx'``.

2. **Credit Karma list** (hard-coded — the user pasted these in PR ARGUMENTS).
   Each row becomes one `ExpectedAccount` with ``source='credit_karma'``,
   ``last_4`` extracted from the parenthesised mask in the institution column.

After seed insert, every new row is auto-linked to a live ``Account`` if its
``last_4`` matches the trailing four characters of an ``Account.account_number``
(case-insensitive, ignoring non-alphanumerics in the comparison).

Modes:

* ``seed --file <xlsx-path>``: dry-run by default; pass ``--apply`` to commit.
* ``confirm``: walk every ``unconfirmed`` row and prompt
  ``"Active or closed? [a/c/s=skip]"``.

Idempotent: the natural-key UNIQUE constraint
``(institution, account_name, last_4)`` rejects duplicate rows. The seeder
catches ``IntegrityError`` per row, rolls back, and continues.

Run:

.. code-block:: bash

   python -m scripts.seed_expected_accounts seed --file <xlsx-path> [--apply]
   python -m scripts.seed_expected_accounts confirm
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from collections.abc import Callable
from dataclasses import dataclass, field
from typing import TYPE_CHECKING

import openpyxl
from sqlalchemy.exc import IntegrityError

from src.db.connection import SessionLocal
from src.models.brokerage import Account
from src.models.enums import AccountType, Broker, Entity
from src.models.history import ExpectedAccount

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)


# ── Constants ────────────────────────────────────────────────────────────────


XLSX_SHEET_NAME = "Account Summary"

# Aggregate / projection rows to skip when reading the sheet.
_SKIP_NAMES: frozenset[str] = frozenset(
    n.lower()
    for n in (
        "Savings",
        "Retirement",
        "Total",
        "College Savings",
        "A Projected",
        "E Projected",
        "Rate",
        "Yearly",
        "Years",
        "Annual",
    )
)


# expected_account.institution → Broker enum value. Used by the confirm
# walkthrough when an active row has no resolved Account row — the seeder
# offers to create one against the matched broker. Phase 4 added the four
# non-broker institutions (FT, NW Mutual, F&G, GSK).
_INSTITUTION_TO_BROKER: dict[str, Broker] = {
    "Vanguard": Broker.VANGUARD,
    "E*TRADE": Broker.ETRADE,
    "Charles Schwab": Broker.SCHWAB,
    "Fidelity Investments": Broker.FIDELITY,
    "Franklin Templeton": Broker.FRANKLIN_TEMPLETON,
    "Northwestern Mutual Investment Services": Broker.NW_MUTUAL,
    "F&G Life": Broker.FG_ANNUITY,
    "GSK": Broker.GSK_PENSION,
}


# Best-guess institution per XLSX label. Falls back to the label itself.
_XLSX_INSTITUTION_HINTS: dict[str, str] = {
    "aiden vanguard": "Vanguard",
    "amazon 401k": "Fidelity",
    "amazon stock": "Charles Schwab",
    "amy ira": "Vanguard",
    "amy roth": "Vanguard",
    "bitcoin": "Bitcoin",
    "charles schwab/td": "Charles Schwab",
    "etrade/sharebuilder": "E*TRADE",
    "emerson coverdale": "Vanguard",
    "emerson vanguard": "Vanguard",
    "gsk pension": "GSK",
    "msft 401k": "Fidelity",
    "microsoft stock": "Charles Schwab",
    "templeton": "Franklin Templeton",
    "travis ira": "Vanguard",
    "travis roth": "Vanguard",
    "vanguard": "Vanguard",
}


# Hard-coded snapshot of the Credit Karma account list. 18 rows.
@dataclass(frozen=True)
class _CKRow:
    account_name: str
    institution: str
    last_4_raw: str  # the parenthesised string, e.g. "...6354", "...X724", "...9-01"


_CREDIT_KARMA_ROWS: list[_CKRow] = [
    _CKRow("Cap 1(-6084)", "E*TRADE", "...6354"),
    _CKRow("Joint Tenant ...724", "Charles Schwab", "...X724"),
    _CKRow(
        "Travis D. Sparks Traditional IRA Brokerage Account 32628019",
        "Vanguard",
        "...8019",
    ),
    _CKRow("AMZN RSU", "Charles Schwab", "...X144"),
    _CKRow(
        "Amy C Sparks Rollover IRA Brokerage Account 65344815", "Vanguard", "...4815"
    ),
    _CKRow("BrokerageLink", "Fidelity Investments", "...3015"),
    _CKRow(
        "Emerson D SparksIndividual 529 College Savings Account, Amy C Sparks20818283901",
        "Vanguard",
        "...3901",
    ),
    _CKRow(
        "Aiden C SparksIndividual 529 College Savings Account, Travis D. Sparks25234130901",
        "Vanguard",
        "...9-01",
    ),
    _CKRow(
        "Aiden C SparksIndividual 529 College Savings Account, Travis D. Sparks25234130901",
        "Vanguard",
        "...0901",
    ),
    _CKRow(
        "Travis D. Sparks Roth IRA Brokerage Account 59309844", "Vanguard", "...9844"
    ),
    _CKRow("Health Savings Account", "Fidelity Investments", "...7012"),
    _CKRow(
        "Amy C Sparks Roth IRA Brokerage Account 70862729", "Vanguard", "...2729"
    ),
    _CKRow("Templeton Growth Fund, Inc. - A", "Franklin Templeton", "...8291"),
    _CKRow("90 LIFE", "Northwestern Mutual Investment Services", "...9215"),
    _CKRow("Individual ...316", "Charles Schwab", "...X316"),
    _CKRow("90 LIFE", "Northwestern Mutual Investment Services", "...5148"),
    _CKRow("90 LIFE", "Northwestern Mutual Investment Services", "...7277"),
    _CKRow("Individual - TOD", "Fidelity Investments", "...7759"),
]


# ── Result dataclass ─────────────────────────────────────────────────────────


@dataclass
class SeedResult:
    """Summary of a seed run."""

    inserted_xlsx: int = 0
    inserted_credit_karma: int = 0
    duplicates_skipped: int = 0
    auto_linked: int = 0
    distinct_xlsx_names: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def total_inserted(self) -> int:
        return self.inserted_xlsx + self.inserted_credit_karma


# ── Helpers ──────────────────────────────────────────────────────────────────


def _strip_last_4(raw: str) -> str | None:
    """Extract a 4-char last_4 from the raw mask (e.g. ``"...6354"``).

    Strategy: drop a leading "..." then keep the rightmost 4 alphanumerics.
    Returns None when nothing usable is left.
    """
    if not raw:
        return None
    cleaned = raw.lstrip(".").strip()
    # Keep alphanumeric + dash so masks like "9-01" survive when chosen
    # explicitly.
    candidate = re.sub(r"[^0-9A-Za-z\-]", "", cleaned)
    if not candidate:
        return None
    # Take the last 4 chars (so "9-01" → "9-01", "X724" → "X724",
    # "6354" → "6354"). We preserve case (X stays X) — Account.account_number
    # comparison is case-insensitive.
    return candidate[-4:] if len(candidate) >= 4 else candidate


def _institution_for_xlsx_name(name: str) -> str:
    return _XLSX_INSTITUTION_HINTS.get(name.strip().lower(), name.strip())


def _read_xlsx_account_names(file_path: str) -> list[str]:
    """Re-enumerate distinct non-aggregate account names from the workbook."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    if XLSX_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Workbook is missing required sheet: {XLSX_SHEET_NAME!r}"
        )
    ws = wb[XLSX_SHEET_NAME]
    seen: set[str] = set()
    out: list[str] = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        cell = row[0]
        if not isinstance(cell, str):
            continue
        s = cell.strip()
        if not s or s.lower() in _SKIP_NAMES:
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    wb.close()
    return out


def _normalize_for_match(value: str) -> tuple[str, str]:
    """Return ``(alnum_lower, digits_only)`` for last-4 / account-number matching."""
    alnum = re.sub(r"[^0-9A-Za-z]", "", value or "").lower()
    digits = re.sub(r"[^0-9]", "", value or "")
    return alnum, digits


def _build_account_match_indexes(
    session: Session,
) -> tuple[dict[str, str], dict[str, str]]:
    """Build two suffix indexes over live ``Account.account_number``:

    1. ``alnum_idx``: last 4 alphanumeric chars (lowercased) → Account.id
    2. ``digits_idx``: all-digit suffix (longest fallback) → Account.id

    On collision, the first encountered wins.
    """
    alnum_idx: dict[str, str] = {}
    digits_idx: dict[str, str] = {}
    for acct in session.query(Account).all():
        alnum, digits = _normalize_for_match(acct.account_number)
        if alnum:
            alnum_idx.setdefault(alnum[-4:], acct.id)
        if digits:
            # Index by trailing 4 digits (predominant CK pattern).
            digits_idx.setdefault(digits[-4:], acct.id)
    return alnum_idx, digits_idx


def _maybe_link(
    expected: ExpectedAccount,
    alnum_idx: dict[str, str],
    digits_idx: dict[str, str],
) -> bool:
    """Set ``resolved_account_id`` on a ``last_4`` match. Returns True if linked.

    Match strategy: try alphanumeric-suffix first (handles ``X724`` →
    account ``724``-suffixed), then fall back to digit-only suffix
    (handles ``9-01`` → account ending ``...0901``).
    """
    if expected.last_4 is None:
        return False
    alnum, digits = _normalize_for_match(expected.last_4)

    if alnum:
        # Try the 4-char alphanumeric tail; if that misses, also try the
        # digit-only tail of the same string (so "X724" → "724" → match).
        target = alnum_idx.get(alnum[-4:]) or (
            digits_idx.get(digits[-4:]) if digits else None
        )
        if target is not None:
            expected.resolved_account_id = target
            return True

    if digits:
        target = digits_idx.get(digits[-4:])
        if target is not None:
            expected.resolved_account_id = target
            return True

    return False


# ── Seeder ───────────────────────────────────────────────────────────────────


def _try_insert(
    session: Session, expected: ExpectedAccount, result: SeedResult, kind: str
) -> bool:
    """Insert one row, handling natural-key UNIQUE collisions. Returns True on
    successful insert."""
    session.add(expected)
    try:
        session.flush()
    except IntegrityError:
        session.rollback()
        result.duplicates_skipped += 1
        return False
    if kind == "xlsx":
        result.inserted_xlsx += 1
    elif kind == "credit_karma":
        result.inserted_credit_karma += 1
    return True


def seed(
    session: Session, xlsx_path: str, apply: bool = False
) -> SeedResult:
    """Build expected_account rows from XLSX + Credit Karma list.

    When ``apply`` is False (default), changes are rolled back at the end
    so the database is untouched. The ``SeedResult`` still reflects what
    *would* have been inserted/linked.
    """
    result = SeedResult()
    xlsx_names = _read_xlsx_account_names(xlsx_path)
    result.distinct_xlsx_names = list(xlsx_names)

    # Build the live-account suffix indexes once. Used for auto-linking.
    alnum_idx, digits_idx = _build_account_match_indexes(session)

    # SQLite treats NULL ≠ NULL inside UNIQUE constraints, so the natural-key
    # constraint won't catch duplicate XLSX rows where ``last_4`` is None.
    # Pre-load a (institution, account_name) set for explicit dedup.
    existing_xlsx_keys: set[tuple[str, str]] = {
        (row.institution, row.account_name)
        for row in session.query(ExpectedAccount)
        .filter(ExpectedAccount.last_4.is_(None))
        .all()
    }

    # Pass 1 — XLSX rows (no last_4 known; they'll get linked manually later).
    for name in xlsx_names:
        institution = _institution_for_xlsx_name(name)
        if (institution, name) in existing_xlsx_keys:
            result.duplicates_skipped += 1
            continue
        e = ExpectedAccount(
            institution=institution,
            account_name=name,
            last_4=None,
            source="xlsx",
            status="unconfirmed",
        )
        if _try_insert(session, e, result, "xlsx"):
            existing_xlsx_keys.add((institution, name))
            if _maybe_link(e, alnum_idx, digits_idx):
                result.auto_linked += 1

    # Pass 2 — Credit Karma rows.
    for ck in _CREDIT_KARMA_ROWS:
        last_4 = _strip_last_4(ck.last_4_raw)
        e = ExpectedAccount(
            institution=ck.institution,
            account_name=ck.account_name,
            last_4=last_4,
            source="credit_karma",
            status="unconfirmed",
        )
        if _try_insert(
            session, e, result, "credit_karma"
        ) and _maybe_link(e, alnum_idx, digits_idx):
            result.auto_linked += 1

    if apply:
        session.commit()
    else:
        session.rollback()

    return result


# ── Interactive confirm ──────────────────────────────────────────────────────


def _offer_account_creation(
    row: ExpectedAccount,
    session: Session,
    input_fn: Callable[[str], str],
    output_fn: Callable[[str], None],
) -> bool:
    """Prompt the operator to create an ``Account`` row for ``row``.

    Returns True if an Account was created and linked; False if the operator
    declined or the institution has no Broker mapping.
    """
    broker = _INSTITUTION_TO_BROKER.get(row.institution)
    if broker is None:
        return False
    prompt = (
        f"  No live Account row for this expected_account.\n"
        f"  Create one as broker={broker.value}? "
        f"Enter account_number (blank to skip): "
    )
    try:
        account_number = input_fn(prompt).strip()
    except EOFError:
        return False
    if not account_number:
        return False
    new_account = Account(
        broker=broker.value,
        account_number=account_number,
        account_name=row.account_name[:128],
        account_type=AccountType.OTHER.value,
        entity=Entity.PERSONAL.value,
    )
    session.add(new_account)
    session.flush()
    row.resolved_account_id = new_account.id
    output_fn(
        f"  Created Account(broker={broker.value},"
        f" number={account_number}); linked.\n"
    )
    return True


def confirm_interactive(
    session: Session,
    *,
    input_fn: Callable[[str], str] = input,
    output_fn: Callable[[str], None] = print,
) -> dict[str, int]:
    """Walk every ``unconfirmed`` ExpectedAccount and prompt the user.

    Response codes: ``a`` → active, ``c`` → closed, ``s`` (or anything else)
    → skip (leave as ``unconfirmed``).

    For rows marked ``a`` whose ``resolved_account_id`` is None and whose
    ``institution`` maps to a known ``Broker``, the operator is additionally
    prompted to create the missing ``Account`` row.
    """
    counts: dict[str, int] = {
        "active": 0, "closed": 0, "skipped": 0, "accounts_created": 0,
    }
    rows = (
        session.query(ExpectedAccount)
        .filter(ExpectedAccount.status == "unconfirmed")
        .order_by(ExpectedAccount.institution, ExpectedAccount.account_name)
        .all()
    )
    if not rows:
        output_fn("No unconfirmed expected_account rows to review.")
        return counts

    for row in rows:
        last_4_disp = f" ...{row.last_4}" if row.last_4 else ""
        prompt = (
            f"[{row.source}] {row.institution} / {row.account_name}{last_4_disp}\n"
            "  Active or closed? [a/c/s=skip]: "
        )
        try:
            answer = input_fn(prompt).strip().lower()
        except EOFError:
            answer = "s"
        if answer == "a":
            row.status = "active"
            counts["active"] += 1
            if row.resolved_account_id is None and _offer_account_creation(
                row, session, input_fn, output_fn
            ):
                counts["accounts_created"] += 1
        elif answer == "c":
            row.status = "closed"
            counts["closed"] += 1
        else:
            counts["skipped"] += 1

    session.commit()
    output_fn(
        f"Confirmed: {counts['active']} active, {counts['closed']} closed, "
        f"{counts['skipped']} skipped, "
        f"{counts['accounts_created']} accounts created."
    )
    return counts


# ── CLI ──────────────────────────────────────────────────────────────────────


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Seed expected_account from XLSX + Credit Karma list."
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_seed = sub.add_parser("seed", help="Seed rows (default: dry-run).")
    p_seed.add_argument(
        "--file",
        required=True,
        help="Path to the Savings & Retirement Plan XLSX.",
    )
    p_seed.add_argument(
        "--apply",
        action="store_true",
        help="Commit changes. Without this flag, runs as a dry-run.",
    )

    sub.add_parser("confirm", help="Interactive walk through unconfirmed rows.")

    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)

    if args.cmd == "seed":
        session = SessionLocal()
        try:
            result = seed(session, xlsx_path=args.file, apply=args.apply)
        finally:
            session.close()
        mode = "APPLY" if args.apply else "DRY-RUN"
        print(f"[{mode}] expected_account seed:")
        print(f"  XLSX inserted:         {result.inserted_xlsx}")
        print(f"  Credit Karma inserted: {result.inserted_credit_karma}")
        print(f"  Duplicates skipped:    {result.duplicates_skipped}")
        print(f"  Auto-linked by last_4: {result.auto_linked}")
        print(f"  Distinct XLSX names:   {len(result.distinct_xlsx_names)}")
        if result.errors:
            print(f"  Errors ({len(result.errors)}):")
            for err in result.errors:
                print(f"    - {err}")
        return 0

    if args.cmd == "confirm":
        session = SessionLocal()
        try:
            confirm_interactive(session)
        finally:
            session.close()
        return 0

    # Should not reach here — required=True on the subparser.
    return 2


if __name__ == "__main__":
    sys.exit(main())
