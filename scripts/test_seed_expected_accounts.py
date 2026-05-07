"""Tests for ``scripts/seed_expected_accounts.py`` (Phase 3 T17)."""

from __future__ import annotations

from collections.abc import Generator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

# Import the module under test (filename uses underscores so we can import directly).
from scripts import seed_expected_accounts as seeder

# Side-effect imports: register all ORM tables on Base.metadata before
# create_all runs.
from src.models import brokerage as _brokerage_models  # noqa: F401
from src.models import history as _history_models  # noqa: F401
from src.models.base import Base
from src.models.brokerage import Account
from src.models.enums import AccountType, Broker, Entity
from src.models.history import ExpectedAccount

# ── Fixtures ─────────────────────────────────────────────────────────────────


@pytest.fixture()
def engine() -> Generator[Any, None, None]:
    e = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(e)
    yield e
    e.dispose()


@pytest.fixture()
def session(engine: Any) -> Generator[Session, None, None]:
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    s = SessionLocal()
    yield s
    s.close()


@pytest.fixture()
def fake_xlsx(tmp_path: Path) -> str:
    """Build a minimal Account Summary sheet that mimics the real workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Summary"
    # Header row: column A is "Account", columns B..D are dates (we don't read
    # them — the seeder only enumerates names).
    ws.cell(row=1, column=1, value="Account")
    # Aggregates / projection rows that must be skipped.
    skip_rows = [
        "Savings",
        "Retirement",
        "Total",
        "College Savings",
        "A Projected",
        "E Projected",
        "Rate",
        "Yearly",
        "Years",
        "Annual",
    ]
    real_rows = [
        "Charles Schwab/TD",
        "ETrade/Sharebuilder",
        "Vanguard",
        "Templeton",
        "Travis Roth",
    ]
    for i, name in enumerate(skip_rows + real_rows, start=2):
        ws.cell(row=i, column=1, value=name)

    path = tmp_path / "fake.xlsx"
    wb.save(str(path))
    return str(path)


def _make_account(
    session: Session,
    *,
    broker: Broker = Broker.VANGUARD,
    account_number: str = "59309844",
    account_type: AccountType = AccountType.TAXABLE,
) -> Account:
    a = Account(
        broker=broker.value,
        account_number=account_number,
        account_type=account_type.value,
        entity=Entity.PERSONAL.value,
        tax_sheltered=False,
    )
    session.add(a)
    session.flush()
    return a


# ── _strip_last_4 / _normalize_for_match ────────────────────────────────────


class TestHelpers:
    def test_strip_last_4_basic(self) -> None:
        assert seeder._strip_last_4("...6354") == "6354"
        assert seeder._strip_last_4("...X724") == "X724"

    def test_strip_last_4_dashed(self) -> None:
        # The "9-01" mask preserves the dash because the 4-char tail is "9-01".
        assert seeder._strip_last_4("...9-01") == "9-01"

    def test_strip_last_4_short(self) -> None:
        # A 3-char input stays 3 chars (we only truncate when >=4).
        assert seeder._strip_last_4("...316") == "316"

    def test_strip_last_4_empty(self) -> None:
        assert seeder._strip_last_4("") is None
        assert seeder._strip_last_4("...") is None


# ── seed() — dry-run / apply / idempotent / auto-link ───────────────────────


class TestSeed:
    def test_dry_run_inserts_nothing(
        self, session: Session, fake_xlsx: str
    ) -> None:
        result = seeder.seed(session, xlsx_path=fake_xlsx, apply=False)
        # SeedResult counts what *would* be inserted...
        assert result.inserted_xlsx == 5
        assert result.inserted_credit_karma == 18
        # ...but the DB is untouched.
        assert session.query(ExpectedAccount).count() == 0

    def test_apply_inserts_both_sources(
        self, session: Session, fake_xlsx: str
    ) -> None:
        result = seeder.seed(session, xlsx_path=fake_xlsx, apply=True)
        assert result.inserted_xlsx == 5
        assert result.inserted_credit_karma == 18

        rows = session.query(ExpectedAccount).all()
        assert len(rows) == 5 + 18

        sources = {r.source for r in rows}
        assert sources == {"xlsx", "credit_karma"}

        statuses = {r.status for r in rows}
        assert statuses == {"unconfirmed"}

    def test_re_running_is_idempotent(
        self, session: Session, fake_xlsx: str
    ) -> None:
        first = seeder.seed(session, xlsx_path=fake_xlsx, apply=True)
        before = session.query(ExpectedAccount).count()

        second = seeder.seed(session, xlsx_path=fake_xlsx, apply=True)
        after = session.query(ExpectedAccount).count()

        assert before == after  # no new rows
        assert second.inserted_xlsx == 0
        assert second.inserted_credit_karma == 0
        assert second.duplicates_skipped == first.total_inserted

    def test_auto_links_by_last_4_to_live_account(
        self, session: Session, fake_xlsx: str
    ) -> None:
        # Pre-seed a Vanguard account whose number ends in 9844 — it should
        # be auto-linked to the Travis Roth IRA Credit Karma row (...9844).
        acct = _make_account(
            session,
            broker=Broker.VANGUARD,
            account_number="59309844",
            account_type=AccountType.ROTH_IRA,
        )
        session.commit()

        result = seeder.seed(session, xlsx_path=fake_xlsx, apply=True)

        linked = (
            session.query(ExpectedAccount)
            .filter(ExpectedAccount.last_4 == "9844")
            .one_or_none()
        )
        assert linked is not None
        assert linked.resolved_account_id == acct.id
        assert result.auto_linked >= 1

    def test_natural_key_collision_drops_one_of_the_dup_pair(
        self, session: Session, fake_xlsx: str
    ) -> None:
        """The 9-01 vs 0901 entries differ in last_4 — both insert. But if a
        duplicate exact (institution, account_name, last_4) existed already,
        the seeder rolls back that one row and continues."""
        existing = ExpectedAccount(
            institution="Vanguard",
            account_name=(
                "Travis D. Sparks Roth IRA Brokerage Account 59309844"
            ),
            last_4="9844",
            source="manual",
            status="active",
        )
        session.add(existing)
        session.commit()

        result = seeder.seed(session, xlsx_path=fake_xlsx, apply=True)

        # That credit_karma row collides with the pre-existing manual row.
        assert result.duplicates_skipped >= 1
        # Existing row's status remains "active" — the seeder did not touch it.
        again = (
            session.query(ExpectedAccount)
            .filter(
                ExpectedAccount.institution == "Vanguard",
                ExpectedAccount.last_4 == "9844",
            )
            .one()
        )
        assert again.status == "active"
        assert again.source == "manual"


# ── confirm_interactive() ────────────────────────────────────────────────────


class TestConfirmInteractive:
    def test_walks_unconfirmed_and_applies_responses(
        self, session: Session
    ) -> None:
        e1 = ExpectedAccount(
            institution="A", account_name="X", source="manual", status="unconfirmed"
        )
        e2 = ExpectedAccount(
            institution="A", account_name="Y", source="manual", status="unconfirmed"
        )
        e3 = ExpectedAccount(
            institution="A", account_name="Z", source="manual", status="unconfirmed"
        )
        session.add_all([e1, e2, e3])
        session.commit()

        # When marked active, institution "A" is not in the broker mapping →
        # no Account creation prompt fires, so all 3 inputs map to status only.
        responses = iter(["a", "c", "s"])
        outputs: list[str] = []
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _prompt: next(responses),
            output_fn=outputs.append,
        )
        assert counts == {
            "active": 1, "closed": 1, "skipped": 1, "accounts_created": 0,
        }

        statuses = {
            row.account_name: row.status
            for row in session.query(ExpectedAccount).all()
        }
        assert statuses == {"X": "active", "Y": "closed", "Z": "unconfirmed"}

    def test_no_unconfirmed_rows_is_a_noop(self, session: Session) -> None:
        outputs: list[str] = []
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _p: "a",
            output_fn=outputs.append,
        )
        assert counts == {
            "active": 0, "closed": 0, "skipped": 0, "accounts_created": 0,
        }
        assert any("No unconfirmed" in s for s in outputs)

    def test_active_with_unmapped_institution_offers_account_creation(
        self, session: Session
    ) -> None:
        """Phase-4 institutions (FT, NW Mutual, F&G, GSK) trigger account-create prompt."""
        from src.models.brokerage import Account

        e = ExpectedAccount(
            institution="Franklin Templeton",
            account_name="Templeton Growth Fund",
            last_4="8291",
            source="manual",
            status="unconfirmed",
        )
        session.add(e)
        session.commit()

        responses = iter(["a", "8291"])
        outputs: list[str] = []
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _p: next(responses),
            output_fn=outputs.append,
        )
        assert counts["active"] == 1
        assert counts["accounts_created"] == 1

        # Account row exists with the right broker.
        accts = (
            session.query(Account)
            .filter(Account.broker == "franklin_templeton")
            .all()
        )
        assert len(accts) == 1
        assert accts[0].account_number == "8291"

        # ExpectedAccount is linked.
        e_reloaded = session.query(ExpectedAccount).one()
        assert e_reloaded.resolved_account_id == accts[0].id

    def test_active_with_blank_account_number_skips_creation(
        self, session: Session
    ) -> None:
        """Operator declining the account-create prompt leaves Account un-made."""
        from src.models.brokerage import Account

        e = ExpectedAccount(
            institution="GSK",
            account_name="Cash Balance Pension",
            source="manual",
            status="unconfirmed",
        )
        session.add(e)
        session.commit()

        responses = iter(["a", ""])  # active, then blank account_number
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _p: next(responses),
            output_fn=lambda _o: None,
        )
        assert counts["active"] == 1
        assert counts["accounts_created"] == 0
        assert (
            session.query(Account).filter(Account.broker == "gsk_pension").count()
            == 0
        )

    def test_active_with_nw_mutual_institution_offers_account_creation(
        self, session: Session
    ) -> None:
        """FIX-N: Northwestern Mutual maps to broker=nw_mutual."""
        from src.models.brokerage import Account

        e = ExpectedAccount(
            institution="Northwestern Mutual Investment Services",
            account_name="90 LIFE",
            last_4="9215",
            source="credit_karma",
            status="unconfirmed",
        )
        session.add(e)
        session.commit()

        responses = iter(["a", "17399215"])  # active, supply account_number
        outputs: list[str] = []
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _p: next(responses),
            output_fn=outputs.append,
        )
        assert counts["active"] == 1
        assert counts["accounts_created"] == 1

        accts = (
            session.query(Account)
            .filter(Account.broker == "nw_mutual")
            .all()
        )
        assert len(accts) == 1
        assert accts[0].account_number == "17399215"

        e_reloaded = session.query(ExpectedAccount).one()
        assert e_reloaded.resolved_account_id == accts[0].id

    def test_active_with_fg_annuity_institution_offers_account_creation(
        self, session: Session
    ) -> None:
        """FIX-N: F&G Life maps to broker=fg_annuity."""
        from src.models.brokerage import Account

        e = ExpectedAccount(
            institution="F&G Life",
            account_name="FG AccumulatorPlus 10",
            last_4="2585",
            source="manual",
            status="unconfirmed",
        )
        session.add(e)
        session.commit()

        responses = iter(["a", "MZ152585"])
        outputs: list[str] = []
        counts = seeder.confirm_interactive(
            session,
            input_fn=lambda _p: next(responses),
            output_fn=outputs.append,
        )
        assert counts["active"] == 1
        assert counts["accounts_created"] == 1

        accts = (
            session.query(Account)
            .filter(Account.broker == "fg_annuity")
            .all()
        )
        assert len(accts) == 1
        assert accts[0].account_number == "MZ152585"

        e_reloaded = session.query(ExpectedAccount).one()
        assert e_reloaded.resolved_account_id == accts[0].id
