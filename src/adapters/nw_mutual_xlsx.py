"""Northwestern Mutual whole-life XLSX importer (Phase 4 — Adapter 3).

Reads the single-sheet ``allAccounts.xlsx`` Northwestern Mutual exports, with
columns ``Insured | Account Number | Net Death Benefit | Annualized Premium |
Last Annual Dividend | Loans | Net Accumulated Value``. Writes one
:class:`AccountBalanceSnapshot` per policy, using ``Net Accumulated Value`` as
the balance.

Key behaviors (matching IDEATION.md and the canonical
``xlsx_savings_plan.py`` pattern):

- Policies whose ``Net Accumulated Value`` is the literal string ``"N/A"``
  (NW Mutual's marker for term-only policies with no cash value) are
  skip-with-warning — appended to ``ImportResult.warnings`` so the operator
  sees them, but the batch continues and they do NOT count as failures.
- Account-id lookup is by ``(broker='nw_mutual', account_number=<policy>)``.
  Missing Account rows produce a per-row error; the adapter does NOT
  auto-create accounts (operator must seed via ``seed_expected_accounts``).
- ``as_of`` defaults to the file's filesystem mtime (UTC date).
  CLI ``--as-of YYYY-MM-DD`` overrides.
- Idempotent: rerunning the same workbook for the same ``as_of`` produces
  zero new rows because ``source_row_hash`` collisions are caught.
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import sys
import traceback
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from src.adapters._shared.ingestion import write_ingestion_log
from src.adapters._shared.money import parse_currency, quantize_balance
from src.models.brokerage import Account
from src.models.enums import Broker, IngestionStatus
from src.models.history import AccountBalanceSnapshot
from src.models.ingestion_log import IngestionLog  # noqa: F401 — re-exported for test compat

logger = logging.getLogger(__name__)


# ── Constants ────────────────────────────────────────────────────────────────

SOURCE_TAG = "nw_mutual_xlsx"
"""Value written to ``account_balance_snapshot.source``."""

ADAPTER_NAME = "nw_mutual_xlsx"
"""Identifier written to ``ingestion_log.source``."""

SHEET_NAME = "Life Insurance"
NW_MUTUAL_BROKER = Broker.NW_MUTUAL.value

# Sentinel string NW Mutual writes for policies with no cash value (term-only).
_NA_MARKERS = frozenset({"N/A", "n/a", "NA", "na", "--", "-"})


# ── Result dataclass ─────────────────────────────────────────────────────────


@dataclass
class ImportResult:
    """Summary of an import run."""

    parsed: int = 0
    """Total policy rows parsed from the workbook."""

    imported: int = 0
    """Newly inserted snapshot rows (apply mode only)."""

    matched: int = 0
    """Rows whose account_id resolved to a live Account."""

    dup_skipped: int = 0
    """Rows skipped because an equivalent snapshot already exists."""

    errors: list[str] = field(default_factory=list)
    """Per-row error strings (genuine failures — not N/A skips)."""

    warnings: list[str] = field(default_factory=list)
    """Per-row warning strings (expected non-fatal skips, e.g. N/A rows)."""

    distinct_accounts: list[str] = field(default_factory=list)
    """Distinct policy numbers observed."""


# ── Parsing ──────────────────────────────────────────────────────────────────


def _is_na(value: object) -> bool:
    """True if a Net Accumulated Value cell is NW Mutual's "no cash value" marker."""
    if value is None:
        return True
    return isinstance(value, str) and value.strip() in _NA_MARKERS


_PARSE_ERROR = object()
"""Sentinel: NAV cell could not be parsed (not an N/A marker, not a valid number)."""


def parse_workbook(path: Path) -> list[dict[str, Any]]:
    """Read the workbook and return one dict per policy row.

    Each dict carries ``policy_number``, ``insured``, and the five money fields
    as ``Decimal`` (``net_accum_value`` is ``None`` when the cell is the N/A
    marker, or the ``_PARSE_ERROR`` sentinel when the value is present but
    unparseable — callers should treat the latter as a per-row error).
    """
    wb = openpyxl.load_workbook(path, data_only=True, keep_links=False)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"workbook missing sheet '{SHEET_NAME}'")
    ws = wb[SHEET_NAME]

    out: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        insured_cell = ws.cell(row=row_idx, column=1).value
        policy_cell = ws.cell(row=row_idx, column=2).value
        if insured_cell is None and policy_cell is None:
            continue
        insured = str(insured_cell).strip() if insured_cell is not None else ""
        policy_number = str(policy_cell).strip() if policy_cell is not None else ""
        if not policy_number:
            continue

        nav_raw = ws.cell(row=row_idx, column=7).value
        if _is_na(nav_raw):
            net_accum_value: Any = None
        else:
            try:
                net_accum_value = parse_currency(nav_raw)
            except (ValueError, Exception):  # noqa: BLE001 — per-row isolation
                net_accum_value = _PARSE_ERROR

        out.append(
            {
                "policy_number": policy_number,
                "insured": insured,
                "net_death_benefit": parse_currency(ws.cell(row=row_idx, column=3).value),
                "annualized_premium": parse_currency(ws.cell(row=row_idx, column=4).value),
                "last_annual_dividend": parse_currency(ws.cell(row=row_idx, column=5).value),
                "loans": parse_currency(ws.cell(row=row_idx, column=6).value),
                "net_accum_value": net_accum_value,
            }
        )
    return out


# ── Hashing ──────────────────────────────────────────────────────────────────


def _row_hash(policy_number: str, as_of: date, balance: Decimal) -> str:
    """SHA256 of the canonical row identity tuple."""
    payload = "|".join(
        (policy_number, as_of.isoformat(), str(quantize_balance(balance)))
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


# ── Import ───────────────────────────────────────────────────────────────────


def import_balances(
    path: Path,
    *,
    dry_run: bool = True,
    session: Session | None = None,
    as_of: date | None = None,
) -> ImportResult:
    """Import NW Mutual policy balances into ``account_balance_snapshot``.

    Args:
        path:    Path to the XLSX workbook.
        dry_run: When True, parse-and-count only; never write to ``session``.
                 Default True to protect the live DB during exploration.
        session: SQLAlchemy session. Required when ``dry_run`` is False.
        as_of:   Snapshot date. Defaults to the file's mtime (UTC date).

    Returns:
        :class:`ImportResult` with counts and per-row errors/warnings.
    """
    result = ImportResult()
    path = Path(path)
    if as_of is None:
        try:
            as_of = datetime.fromtimestamp(path.stat().st_mtime, tz=UTC).date()
        except OSError as exc:
            result.errors.append(f"stat failed for {path.name}: {exc}")
            if session is not None and not dry_run:
                write_ingestion_log(
                    session,
                    source=ADAPTER_NAME,
                    records_processed=0,
                    records_failed=1,
                    status=IngestionStatus.FAILURE,
                    error_detail=result.errors[-1],
                )
            return result

    try:
        rows = parse_workbook(path)
    except Exception as exc:  # noqa: BLE001 — surface workbook-level failures
        result.errors.append(f"parse_workbook failed: {exc}")
        if session is not None and not dry_run:
            write_ingestion_log(
                session,
                source=ADAPTER_NAME,
                records_processed=0,
                records_failed=1,
                status=IngestionStatus.FAILURE,
                error_detail=result.errors[-1],
            )
        return result

    result.parsed = len(rows)
    result.distinct_accounts = sorted({r["policy_number"] for r in rows})

    # Partition into writable rows, N/A skips (warnings), and parse errors.
    writable: list[dict[str, Any]] = []
    for row in rows:
        if row["net_accum_value"] is None:
            # N/A is expected for term-only policies — goes to warnings only.
            # Strip insured name to avoid PII leak via /api/health.
            result.warnings.append(
                f"policy {row['policy_number']}: "
                "Net Accumulated Value is N/A — skipped"
            )
            continue
        if row["net_accum_value"] is _PARSE_ERROR:
            # Unparseable non-N/A value — this is a genuine error.
            result.errors.append(
                f"policy {row['policy_number']}: "
                "Net Accumulated Value could not be parsed"
            )
            continue
        writable.append(row)

    if dry_run:
        return result

    if session is None:
        result.errors.append("session required when dry_run=False")
        return result

    # ── Apply ────────────────────────────────────────────────────────────────
    for row in writable:
        policy_number = row["policy_number"]
        insured = row["insured"]
        balance: Decimal = row["net_accum_value"]
        raw_account_name = f"NW Mutual {insured} {policy_number}".strip()
        record_label = f"{raw_account_name}@{as_of.isoformat()}"
        row_hash = _row_hash(policy_number, as_of, balance)

        # Idempotency check by the dedup hash.
        existing = (
            session.query(AccountBalanceSnapshot.id)
            .filter(AccountBalanceSnapshot.source_row_hash == row_hash)
            .first()
        )
        if existing is not None:
            result.dup_skipped += 1
            continue

        # Resolve the Account row (NW Mutual broker, account_number == policy).
        account = (
            session.query(Account)
            .filter(
                Account.broker == NW_MUTUAL_BROKER,
                Account.account_number == policy_number,
            )
            .first()
        )
        if account is None:
            result.errors.append(
                f"policy {policy_number}: no Account row "
                f"(broker={NW_MUTUAL_BROKER}, account_number={policy_number})"
            )
            continue

        try:
            with session.begin_nested():
                session.add(
                    AccountBalanceSnapshot(
                        account_id=account.id,
                        raw_account_name=raw_account_name,
                        as_of=as_of,
                        balance=balance,
                        source=SOURCE_TAG,
                        source_row_hash=row_hash,
                    )
                )
            result.imported += 1
            result.matched += 1
        except IntegrityError:
            # Natural-key UNIQUE collision — same logical row already exists.
            result.dup_skipped += 1
        except Exception as exc:  # noqa: BLE001 — per-record isolation
            result.errors.append(f"{record_label}: {exc}")
            logger.warning("nw_mutual_xlsx: row %s failed: %s",
                           record_label, exc, exc_info=True)

    session.commit()

    # ── Audit log ────────────────────────────────────────────────────────────
    # Only genuine errors drive the status; N/A skips in warnings are expected.
    status = (
        IngestionStatus.SUCCESS
        if not result.errors
        else IngestionStatus.PARTIAL_FAILURE
    )
    # Include warnings in error_detail for visibility, but only errors fail.
    detail_parts: list[str] = []
    if result.warnings:
        detail_parts.append("warnings:\n" + "\n".join(result.warnings))
    if result.errors:
        detail_parts.append("errors:\n" + "\n".join(result.errors))
    write_ingestion_log(
        session,
        source=ADAPTER_NAME,
        records_processed=result.imported + result.dup_skipped,
        records_failed=len(result.errors),
        status=status,
        error_detail="\n".join(detail_parts) or None,
    )
    return result


# ── CLI ──────────────────────────────────────────────────────────────────────


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="python -m src.adapters.nw_mutual_xlsx",
        description="Import Northwestern Mutual whole-life balances from XLSX.",
    )
    sub = p.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser(
        "import-balances",
        help="Read allAccounts.xlsx and import policy balance snapshots.",
    )
    s.add_argument("--file", required=True, help="Path to the XLSX workbook.")
    s.add_argument(
        "--apply",
        action="store_true",
        help="Actually write to the live DB. Default is dry-run.",
    )
    s.add_argument(
        "--as-of",
        dest="as_of",
        default=None,
        help="Snapshot date YYYY-MM-DD. Defaults to file mtime.",
    )
    return p


def _print_summary(result: ImportResult, dry_run: bool) -> None:
    mode = "DRY RUN" if dry_run else "APPLIED"
    print(f"=== nw_mutual_xlsx ({mode}) ===")
    print(f"  parsed       : {result.parsed}")
    print(f"  imported     : {result.imported}")
    print(f"  dup_skipped  : {result.dup_skipped}")
    print(f"  warnings     : {len(result.warnings)}")
    print(f"  errors       : {len(result.errors)}")
    if result.warnings:
        print("  --- warning detail ---")
        for w in result.warnings[:10]:
            print(f"    ~ {w}")
        if len(result.warnings) > 10:
            print(f"    ... {len(result.warnings) - 10} more")
    if result.errors:
        print("  --- error detail ---")
        for e in result.errors[:20]:
            print(f"    * {e}")
        if len(result.errors) > 20:
            print(f"    ... {len(result.errors) - 20} more")


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    if args.cmd != "import-balances":
        return 2

    file_path = Path(args.file)
    as_of: date | None = None
    if args.as_of is not None:
        as_of = date.fromisoformat(args.as_of)

    dry_run = not args.apply

    if dry_run:
        result = import_balances(file_path, dry_run=True, as_of=as_of)
        _print_summary(result, dry_run=True)
        return 0

    try:
        from src.db.connection import get_session  # late import to keep tests light
    except ImportError as exc:  # pragma: no cover — environmental
        print(f"cannot import DB session factory: {exc}", file=sys.stderr)
        return 1

    with get_session() as session:
        result = import_balances(
            file_path, dry_run=False, session=session, as_of=as_of
        )
    _print_summary(result, dry_run=False)
    return 0


if __name__ == "__main__":  # pragma: no cover — exercised via CLI
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise SystemExit(1) from None


__all__ = [
    "ADAPTER_NAME",
    "ImportResult",
    "NW_MUTUAL_BROKER",
    "SOURCE_TAG",
    "import_balances",
    "parse_workbook",
]
