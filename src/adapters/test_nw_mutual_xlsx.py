"""Tests for the Northwestern Mutual whole-life XLSX importer (Phase 4 T4).

Builds an in-memory fixture workbook that mirrors the real
``nw-mutual/allAccounts.xlsx`` (4 policy rows; the 4th has
``Net Accumulated Value = "N/A"`` and must be skipped with a warning).

Uses an in-memory SQLite engine with FK enforcement, mirroring
``test_xlsx_savings_plan.py``.
"""

from __future__ import annotations

import contextlib
from collections.abc import Generator
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import Session

from src.adapters.nw_mutual_xlsx import (
    ADAPTER_NAME,
    NW_MUTUAL_BROKER,
    SOURCE_TAG,
    ImportResult,
    import_balances,
    parse_workbook,
)
from src.models.base import Base
from src.models.brokerage import Account
from src.models.history import AccountBalanceSnapshot
from src.models.ingestion_log import IngestionLog

# ── Fixtures ─────────────────────────────────────────────────────────────────


@pytest.fixture
def session() -> Generator[Session, None, None]:
    """Fresh in-memory SQLite with FK enforcement. Yields and cleans up."""
    engine = create_engine("sqlite:///:memory:")

    @event.listens_for(engine, "connect")
    def _fk(dbapi_conn, _):  # type: ignore[no-untyped-def]
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    s = Session(bind=engine)
    try:
        yield s
    finally:
        s.close()
        engine.dispose()


def _build_fixture_workbook(path: Path) -> None:
    """Mirror the real allAccounts.xlsx — 4 policy rows, last has N/A balance."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Life Insurance"

    headers = [
        "Insured",
        "Account Number",
        "Net Death Benefit",
        "Annualized Premium",
        "Last Annual Dividend",
        "Loans",
        "Net Accumulated Value",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    rows = [
        ("Aiden C Sparks", "17397277", "$32,216.00", "$200.16", "$80.37", "$0.00", "$3,250.34"),
        ("Travis D Sparks", "18305148", "$28,547.00", "$327.84", "$157.65", "$0.00", "$5,621.63"),
        ("Travis D Sparks", "17399215", "$28,327.00", "$349.80", "$163.75", "$0.00", "$7,280.48"),
        ("Travis D Sparks", "17399232", "$275,000.00", "$601.92", "$117.86", "$0.00", "N/A"),
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)


def _build_invalid_row_workbook(path: Path) -> None:
    """Fixture workbook with one bad NAV cell ($invalid) + 3 valid rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Life Insurance"

    headers = [
        "Insured", "Account Number", "Net Death Benefit", "Annualized Premium",
        "Last Annual Dividend", "Loans", "Net Accumulated Value",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    rows = [
        ("Aiden C Sparks", "17397277", "$32,216.00", "$200.16", "$80.37", "$0.00", "$3,250.34"),
        ("Travis D Sparks", "18305148", "$28,547.00", "$327.84", "$157.65", "$0.00", "$invalid"),
        ("Travis D Sparks", "17399215", "$28,327.00", "$349.80", "$163.75", "$0.00", "$7,280.48"),
        ("Travis D Sparks", "17399232", "$275,000.00", "$601.92", "$117.86", "$0.00", "$1,234.56"),
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)


@pytest.fixture
def fixture_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "nw_mutual_fixture.xlsx"
    _build_fixture_workbook(p)
    return p


def _seed_accounts(session: Session, policies: list[str]) -> dict[str, str]:
    """Seed Account rows for each policy. Returns {policy_number: account_id}."""
    ids: dict[str, str] = {}
    for policy in policies:
        acct = Account(
            broker=NW_MUTUAL_BROKER,
            account_number=policy,
            account_type="other",
            entity="personal",
            account_name=f"NW Mutual {policy}",
        )
        session.add(acct)
        session.flush()
        ids[policy] = acct.id
    session.commit()
    return ids


ALL_POLICIES = ["17397277", "18305148", "17399215", "17399232"]


# ── parse_workbook ───────────────────────────────────────────────────────────


def test_parse_workbook_returns_4_dicts(fixture_xlsx: Path) -> None:
    rows = parse_workbook(fixture_xlsx)
    assert len(rows) == 4
    assert {r["policy_number"] for r in rows} == set(ALL_POLICIES)
    # Required keys per row.
    expected_keys = {
        "policy_number",
        "insured",
        "net_death_benefit",
        "annualized_premium",
        "last_annual_dividend",
        "loans",
        "net_accum_value",
    }
    for r in rows:
        assert expected_keys.issubset(r.keys())


def test_parse_workbook_na_row_yields_none(fixture_xlsx: Path) -> None:
    rows = parse_workbook(fixture_xlsx)
    by_policy = {r["policy_number"]: r for r in rows}
    assert by_policy["17399232"]["net_accum_value"] is None
    # Other rows have parsed Decimals.
    assert by_policy["17397277"]["net_accum_value"] == Decimal("3250.34")
    assert by_policy["18305148"]["net_accum_value"] == Decimal("5621.63")
    assert by_policy["17399215"]["net_accum_value"] == Decimal("7280.48")


def test_parse_workbook_other_columns(fixture_xlsx: Path) -> None:
    rows = parse_workbook(fixture_xlsx)
    aiden = next(r for r in rows if r["policy_number"] == "17397277")
    assert aiden["insured"] == "Aiden C Sparks"
    assert aiden["net_death_benefit"] == Decimal("32216.00")
    assert aiden["annualized_premium"] == Decimal("200.16")
    assert aiden["last_annual_dividend"] == Decimal("80.37")
    assert aiden["loans"] == Decimal("0.00")


# ── import_balances dry-run ──────────────────────────────────────────────────


def test_dry_run_reports_3_inserts_and_skip_warning(
    session: Session, fixture_xlsx: Path
) -> None:
    result = import_balances(fixture_xlsx, dry_run=True, session=session,
                             as_of=date(2026, 5, 7))
    assert isinstance(result, ImportResult)
    # 4 parsed, N/A skip goes to warnings (not errors).
    assert result.parsed == 4
    # N/A row goes to warnings, not errors.
    assert any("17399232" in w and "N/A" in w for w in result.warnings), result.warnings
    # No genuine errors.
    assert result.errors == []
    # DB must be untouched.
    assert session.query(AccountBalanceSnapshot).count() == 0
    assert session.query(IngestionLog).count() == 0


# ── import_balances apply ───────────────────────────────────────────────────


def test_apply_writes_three_snapshots(session: Session, fixture_xlsx: Path) -> None:
    _seed_accounts(session, ALL_POLICIES)
    result = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    assert result.imported == 3
    snaps = session.query(AccountBalanceSnapshot).all()
    assert len(snaps) == 3
    # Each surviving row has the right tag/source/balance.
    by_raw = {s.raw_account_name: s for s in snaps}
    assert "NW Mutual Aiden C Sparks 17397277" in by_raw
    assert by_raw["NW Mutual Aiden C Sparks 17397277"].balance == Decimal("3250.34")
    assert by_raw["NW Mutual Aiden C Sparks 17397277"].source == SOURCE_TAG
    # Account-id linkage by (broker='nw_mutual', account_number).
    seeded = {a.account_number: a.id for a in session.query(Account).all()}
    assert by_raw["NW Mutual Aiden C Sparks 17397277"].account_id == seeded["17397277"]
    # IngestionLog written.
    logs = session.query(IngestionLog).all()
    assert len(logs) == 1
    # Assert the adapter writes using ADAPTER_NAME (not SOURCE_TAG) as the
    # log source — they happen to be equal today but ADAPTER_NAME is the
    # canonical constant for ingestion_log.source.
    assert logs[0].source == ADAPTER_NAME
    assert logs[0].records_processed == 3


def test_apply_is_idempotent(session: Session, fixture_xlsx: Path) -> None:
    _seed_accounts(session, ALL_POLICIES)
    first = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    assert first.imported == 3
    second = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    assert second.imported == 0
    assert second.dup_skipped == 3
    assert session.query(AccountBalanceSnapshot).count() == 3


def test_unmapped_policy_appends_error(session: Session, fixture_xlsx: Path) -> None:
    # Seed only 2 of the 3 valid policies; one will be unmapped.
    _seed_accounts(session, ["17397277", "18305148"])
    result = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    # 2 imported, 1 unmapped error, 1 N/A skip warning.
    assert result.imported == 2
    assert any("17399215" in e for e in result.errors), result.errors
    # N/A skip in warnings, not errors.
    assert any("17399232" in w for w in result.warnings), result.warnings
    snaps = session.query(AccountBalanceSnapshot).all()
    assert len(snaps) == 2


def test_default_as_of_uses_file_mtime(session: Session, fixture_xlsx: Path) -> None:
    """When as_of is omitted, the file mtime date is used."""
    _seed_accounts(session, ALL_POLICIES)
    result = import_balances(fixture_xlsx, dry_run=False, session=session)
    assert result.imported == 3
    from datetime import UTC, datetime
    expected = datetime.fromtimestamp(fixture_xlsx.stat().st_mtime, tz=UTC).date()
    snaps = session.query(AccountBalanceSnapshot).all()
    for s in snaps:
        assert s.as_of == expected


def test_na_skip_goes_to_warnings_not_errors(
    session: Session, fixture_xlsx: Path
) -> None:
    """N/A rows must produce a warning, not an error — SUCCESS status, not PARTIAL."""
    _seed_accounts(session, ALL_POLICIES)
    result = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    assert result.imported == 3
    # N/A skip in warnings only.
    assert any("17399232" in w for w in result.warnings)
    # No genuine errors.
    assert result.errors == []
    # IngestionLog status == "success" (not partial_failure).
    log = session.query(IngestionLog).one()
    assert log.status == "success"
    assert log.records_failed == 0


def test_na_warning_does_not_contain_insured_name(
    session: Session, fixture_xlsx: Path
) -> None:
    """FIX-E: insured name must NOT appear in warnings (PII leak prevention)."""
    _seed_accounts(session, ALL_POLICIES)
    result = import_balances(
        fixture_xlsx, dry_run=False, session=session, as_of=date(2026, 5, 7)
    )
    for w in result.warnings:
        assert "Travis D Sparks" not in w, f"PII in warning: {w}"
        assert "Aiden C Sparks" not in w, f"PII in warning: {w}"


# ── FIX-M: per-row error isolation ──────────────────────────────────────────


def test_per_row_error_isolation_does_not_break_batch(
    tmp_path: Path, session: Session
) -> None:
    """One invalid NAV cell ($invalid) errors; other 3 valid rows still imported."""
    xlsx = tmp_path / "invalid_row.xlsx"
    _build_invalid_row_workbook(xlsx)

    # Seed all 4 policy accounts (17397277, 18305148, 17399215, 17399232).
    _seed_accounts(session, ["17397277", "18305148", "17399215", "17399232"])

    result = import_balances(xlsx, dry_run=False, session=session,
                             as_of=date(2026, 5, 7))

    # 18305148 has "$invalid" → 1 parse error.
    assert len(result.errors) == 1, f"expected 1 error, got: {result.errors}"
    # The other 3 valid rows should be imported.
    assert result.imported == 3, f"expected 3 imported, got: {result.imported}"
    assert session.query(AccountBalanceSnapshot).count() == 3


# ── FIX-3: error strings must not contain insured name (PII) ────────────────


def _build_pii_test_workbook(path: Path) -> None:
    """Workbook with a single policy row owned by 'TestSubject McTester'.

    Used to verify that per-row exception strings in result.errors do not
    contain the insured name — preventing PII leakage into IngestionLog.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Life Insurance"

    headers = [
        "Insured", "Account Number", "Net Death Benefit", "Annualized Premium",
        "Last Annual Dividend", "Loans", "Net Accumulated Value",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # One valid-looking row with a distinctive insured name.
    ws.cell(row=2, column=1, value="TestSubject McTester")
    ws.cell(row=2, column=2, value="99990001")
    ws.cell(row=2, column=3, value="$100,000.00")
    ws.cell(row=2, column=4, value="$500.00")
    ws.cell(row=2, column=5, value="$50.00")
    ws.cell(row=2, column=6, value="$0.00")
    ws.cell(row=2, column=7, value="$12,345.00")

    wb.save(path)


def test_error_string_does_not_contain_insured_name(
    tmp_path: Path, session: Session
) -> None:
    """FIX-3: when a row insert raises, result.errors must NOT contain the
    insured name — PII must not flow into IngestionLog.error_detail."""
    xlsx = tmp_path / "pii_test.xlsx"
    _build_pii_test_workbook(xlsx)

    # Seed the matching Account row so account lookup succeeds.
    acct = Account(
        broker=NW_MUTUAL_BROKER,
        account_number="99990001",
        account_type="other",
        entity="personal",
    )
    session.add(acct)
    session.commit()

    # Patch begin_nested to raise a RuntimeError so the per-row except branch fires.
    @contextlib.contextmanager
    def _failing_begin_nested():  # type: ignore[no-untyped-def]
        raise RuntimeError("injected failure for PII test")
        yield  # pragma: no cover — generator protocol requires a yield

    with patch.object(session, "begin_nested", _failing_begin_nested):
        result = import_balances(xlsx, dry_run=False, session=session,
                                 as_of=date(2026, 5, 7))

    # A per-row error must have been recorded.
    assert len(result.errors) >= 1, f"expected at least 1 error, got: {result.errors}"
    # The insured name must NOT appear in any error string.
    for err in result.errors:
        assert "TestSubject McTester" not in err, (
            f"PII (insured name) found in error string: {err!r}"
        )
