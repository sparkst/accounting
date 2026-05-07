"""Tests for the XLSX 'Savings & Retirement Plan' importer.

These tests build a tiny in-memory workbook fixture so they don't depend on
the user's real spreadsheet. They use an in-memory SQLite engine with FK
enforcement, mirroring ``src/models/test_history_models.py``.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import create_engine, event
from sqlalchemy.orm import Session

from src.adapters.xlsx_savings_plan import (
    PRICES_SOURCE_TAG,
    SB_LOTS_SOURCE_TAG,
    SB_RAW_ACCOUNT_NAME,
    SOURCE_TAG,
    TD_LOTS_SOURCE_TAG,
    TD_RAW_ACCOUNT_NAME,
    ImportResult,
    import_account_balances,
    import_cost_basis_lots,
    import_historical_prices,
)
from src.models.base import Base
from src.models.brokerage import Account  # noqa: F401 — register FK target on metadata
from src.models.history import AccountBalanceSnapshot, CostBasisLot, HistoricalPrice
from src.models.ingestion_log import IngestionLog

# ── Fixtures ─────────────────────────────────────────────────────────────────


@pytest.fixture
def session() -> Session:
    """Fresh in-memory SQLite with FK enforcement (same pattern as test_history_models)."""
    engine = create_engine("sqlite:///:memory:")

    @event.listens_for(engine, "connect")
    def _fk(dbapi_conn, _):  # type: ignore[no-untyped-def]
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine)
    return Session(bind=engine)


def _build_fixture_workbook(path: Path) -> None:
    """Two child accounts × three dates, plus a 'Savings' aggregate row to skip.

    Layout (headers in row 1):

        Account            | 2024-12-11 | 2024-06-27 | 2017-08-08
        Savings            | 100        | 200        | 300        ← skip
          Charles Schwab/TD| 60         | 120        | 180
          Vanguard         | 40         | 80         | 120.50
          (blank row)
          Bitcoin          | None       | '#N/A'     | 0
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Account Summary"

    # Header row.
    ws.cell(row=1, column=1, value="Account")
    ws.cell(row=1, column=2, value=datetime(2024, 12, 11))
    ws.cell(row=1, column=3, value=datetime(2024, 6, 27))
    ws.cell(row=1, column=4, value=datetime(2017, 8, 8))

    # Aggregate row (must be skipped).
    ws.cell(row=2, column=1, value="Savings")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=2, column=3, value=200)
    ws.cell(row=2, column=4, value=300)

    # Real account 1 (note the leading-spaces indent — must be stripped).
    ws.cell(row=3, column=1, value="  Charles Schwab/TD")
    ws.cell(row=3, column=2, value=60)
    ws.cell(row=3, column=3, value=120)
    ws.cell(row=3, column=4, value=180)

    # Real account 2 — non-integer balances to verify Decimal precision.
    ws.cell(row=4, column=1, value="  Vanguard")
    ws.cell(row=4, column=2, value=40.25)
    ws.cell(row=4, column=3, value=80.10)
    ws.cell(row=4, column=4, value=120.50)

    # Blank account-name row → ignored.
    ws.cell(row=5, column=1, value=None)

    # Account with mix of None and #N/A — only the numeric cell should land.
    ws.cell(row=6, column=1, value="  Bitcoin")
    ws.cell(row=6, column=2, value=None)
    ws.cell(row=6, column=3, value="#N/A")
    ws.cell(row=6, column=4, value=0)

    wb.save(path)


@pytest.fixture
def fixture_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "savings_fixture.xlsx"
    _build_fixture_workbook(p)
    return p


# ── Tests ────────────────────────────────────────────────────────────────────


def test_imports_expected_row_count(session: Session, fixture_xlsx: Path) -> None:
    """Three accounts × three dates = 9 cells, but Bitcoin only has 1 numeric.

    Schwab: 3 dates → 3 rows.
    Vanguard: 3 dates → 3 rows.
    Bitcoin: only the third date is numeric (0) → 1 row.
    Savings: aggregate, skipped entirely.
    Expected: 7 rows.
    """
    result = import_account_balances(str(fixture_xlsx), dry_run=False, session=session)

    assert isinstance(result, ImportResult)
    assert result.imported == 7, f"unexpected import count: {result}"
    assert result.dup_skipped == 0
    assert result.errors == []
    assert session.query(AccountBalanceSnapshot).count() == 7


def test_aggregate_savings_row_is_skipped(session: Session, fixture_xlsx: Path) -> None:
    import_account_balances(str(fixture_xlsx), dry_run=False, session=session)

    names = {
        s.raw_account_name
        for s in session.query(AccountBalanceSnapshot).all()
    }
    assert "Savings" not in names
    assert "Charles Schwab/TD" in names
    assert "Vanguard" in names
    assert "Bitcoin" in names


def test_rerun_is_idempotent(session: Session, fixture_xlsx: Path) -> None:
    """Re-running over the same workbook produces zero new rows; everything is dup_skipped."""
    first = import_account_balances(str(fixture_xlsx), dry_run=False, session=session)
    assert first.imported == 7

    second = import_account_balances(str(fixture_xlsx), dry_run=False, session=session)
    assert second.imported == 0
    assert second.dup_skipped == 7
    assert session.query(AccountBalanceSnapshot).count() == 7


def test_decimal_precision_preserved(session: Session, fixture_xlsx: Path) -> None:
    """Vanguard's 40.25, 80.10, 120.50 must land as Decimal with two places intact."""
    import_account_balances(str(fixture_xlsx), dry_run=False, session=session)

    vanguard_rows = (
        session.query(AccountBalanceSnapshot)
        .filter(AccountBalanceSnapshot.raw_account_name == "Vanguard")
        .order_by(AccountBalanceSnapshot.as_of)
        .all()
    )
    assert len(vanguard_rows) == 3
    balances = [r.balance for r in vanguard_rows]
    # All Decimals, not floats.
    assert all(isinstance(b, Decimal) for b in balances)
    # The DB column is Numeric(14,2). Values quantize to 2 dp; sum stays exact.
    assert sum(balances) == Decimal("240.85")


def test_unmatched_account_id_is_null(session: Session, fixture_xlsx: Path) -> None:
    """First-pass importer leaves account_id NULL but populates raw_account_name."""
    result = import_account_balances(str(fixture_xlsx), dry_run=False, session=session)

    rows = session.query(AccountBalanceSnapshot).all()
    assert all(r.account_id is None for r in rows)
    assert all(r.raw_account_name for r in rows)
    assert all(r.source == SOURCE_TAG for r in rows)
    assert result.matched == 0
    assert result.unmatched == 7


def test_dry_run_writes_nothing(session: Session, fixture_xlsx: Path) -> None:
    """Dry-run must not insert anything but should still enumerate accounts."""
    result = import_account_balances(str(fixture_xlsx), dry_run=True, session=None)

    assert result.imported == 0
    assert result.dup_skipped == 0
    # 3 distinct non-aggregate names: Charles Schwab/TD, Vanguard, Bitcoin.
    assert set(result.distinct_accounts) == {"Charles Schwab/TD", "Vanguard", "Bitcoin"}
    assert session.query(AccountBalanceSnapshot).count() == 0


def test_ingestion_log_written_on_apply(session: Session, fixture_xlsx: Path) -> None:
    """A successful apply run records exactly one IngestionLog row."""
    import_account_balances(str(fixture_xlsx), dry_run=False, session=session)

    logs = session.query(IngestionLog).all()
    assert len(logs) == 1
    assert logs[0].source == "xlsx_savings_plan"
    assert logs[0].records_processed == 7
    assert logs[0].records_failed == 0


# ── T11 — Historical Prices ──────────────────────────────────────────────────


def _build_prices_workbook(path: Path) -> None:
    """Mimic the real 'Historical Prices' sheet structure.

    Row 1: bucket labels (ignored)
    Row 2: original (unadjusted) dates (ignored)
    Row 3: 'Symbol' + four weekend-adjusted dates → these are trade_date
    Row 4+: symbol rows. Mix in '#N/A' and None to verify skip behaviour.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Historical Prices"

    # Row 1 — bucket labels.
    ws.cell(row=1, column=1, value=None)
    ws.cell(row=1, column=2, value="Yesterday")
    ws.cell(row=1, column=3, value="30 Days")
    ws.cell(row=1, column=4, value="90 days")
    ws.cell(row=1, column=5, value="1 year")

    # Row 2 — "original" dates (ignored by importer).
    ws.cell(row=2, column=1, value="Original Date used for row 3")
    ws.cell(row=2, column=2, value=datetime(2026, 5, 2))
    ws.cell(row=2, column=3, value=datetime(2026, 4, 3))
    ws.cell(row=2, column=4, value=datetime(2026, 2, 2))
    ws.cell(row=2, column=5, value=datetime(2025, 5, 3))

    # Row 3 — header. trade_dates live here.
    ws.cell(row=3, column=1, value="Symbol")
    ws.cell(row=3, column=2, value=datetime(2026, 5, 1))
    ws.cell(row=3, column=3, value=datetime(2026, 4, 3))
    ws.cell(row=3, column=4, value=datetime(2026, 2, 2))
    ws.cell(row=3, column=5, value=datetime(2025, 5, 2))

    # Two real symbols × four dates.
    ws.cell(row=4, column=1, value="AMZN")
    ws.cell(row=4, column=2, value=268.26)
    ws.cell(row=4, column=3, value=212.79)
    ws.cell(row=4, column=4, value=242.96)
    ws.cell(row=4, column=5, value=189.98)

    ws.cell(row=5, column=1, value="ARR")
    ws.cell(row=5, column=2, value=17.73)
    ws.cell(row=5, column=3, value=17.34)
    ws.cell(row=5, column=4, value=17.06)
    ws.cell(row=5, column=5, value=16.22)

    wb.save(path)


def _build_prices_workbook_with_na(path: Path) -> None:
    """Same shape as above but with #N/A and blank cells sprinkled in."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Historical Prices"

    ws.cell(row=1, column=1, value=None)
    ws.cell(row=1, column=2, value="Yesterday")
    ws.cell(row=1, column=3, value="30 Days")

    ws.cell(row=2, column=1, value="orig")
    ws.cell(row=2, column=2, value=datetime(2026, 5, 2))
    ws.cell(row=2, column=3, value=datetime(2026, 4, 3))

    ws.cell(row=3, column=1, value="Symbol")
    ws.cell(row=3, column=2, value=datetime(2026, 5, 1))
    ws.cell(row=3, column=3, value=datetime(2026, 4, 3))

    # AMZN: one good, one '#N/A'.
    ws.cell(row=4, column=1, value="AMZN")
    ws.cell(row=4, column=2, value=268.26)
    ws.cell(row=4, column=3, value="#N/A")

    # GOOG: one None, one good.
    ws.cell(row=5, column=1, value="GOOG")
    ws.cell(row=5, column=2, value=None)
    ws.cell(row=5, column=3, value=150.5)

    # Empty symbol row — entirely ignored.
    ws.cell(row=6, column=1, value=None)

    wb.save(path)


@pytest.fixture
def prices_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "prices_fixture.xlsx"
    _build_prices_workbook(p)
    return p


@pytest.fixture
def prices_xlsx_na(tmp_path: Path) -> Path:
    p = tmp_path / "prices_na_fixture.xlsx"
    _build_prices_workbook_with_na(p)
    return p


def test_import_historical_prices_round_trip(
    session: Session, prices_xlsx: Path
) -> None:
    """2 symbols × 4 weekend-adjusted dates = 8 rows."""
    result = import_historical_prices(str(prices_xlsx), dry_run=False, session=session)

    assert isinstance(result, ImportResult)
    assert result.imported == 8, f"unexpected import count: {result}"
    assert result.dup_skipped == 0
    assert result.errors == []

    rows = session.query(HistoricalPrice).all()
    assert len(rows) == 8
    assert {r.symbol for r in rows} == {"AMZN", "ARR"}
    assert all(r.source == PRICES_SOURCE_TAG for r in rows)
    # Trade-date should match row 3 (weekend-adjusted), not row 2.
    amzn_dates = {r.trade_date for r in rows if r.symbol == "AMZN"}
    assert amzn_dates == {
        date(2026, 5, 1),
        date(2026, 4, 3),
        date(2026, 2, 2),
        date(2025, 5, 2),
    }


def test_import_historical_prices_skips_na(
    session: Session, prices_xlsx_na: Path
) -> None:
    """'#N/A' and None cells must not produce HistoricalPrice rows."""
    result = import_historical_prices(
        str(prices_xlsx_na), dry_run=False, session=session
    )

    # AMZN: 1 good + 1 #N/A → 1
    # GOOG: 1 None + 1 good → 1
    # Total: 2
    assert result.imported == 2, f"unexpected count: {result}"
    rows = session.query(HistoricalPrice).all()
    assert len(rows) == 2
    by_symbol = {r.symbol: r.trade_date for r in rows}
    assert by_symbol == {
        "AMZN": date(2026, 5, 1),
        "GOOG": date(2026, 4, 3),
    }


def test_import_historical_prices_dedup(
    session: Session, prices_xlsx: Path
) -> None:
    """Re-running the same workbook must produce 0 new rows."""
    first = import_historical_prices(
        str(prices_xlsx), dry_run=False, session=session
    )
    assert first.imported == 8

    second = import_historical_prices(
        str(prices_xlsx), dry_run=False, session=session
    )
    assert second.imported == 0
    assert second.dup_skipped == 8
    assert session.query(HistoricalPrice).count() == 8


# ── T16 — Cost Basis Lots ────────────────────────────────────────────────────


def _build_lots_workbook(
    path: Path,
    *,
    include_td: bool = True,
    include_sb: bool = True,
    include_na_rows: bool = False,
) -> None:
    """Build a workbook with TD GainLoss Raw and/or SB Raw sheets.

    TD layout:  header row 2, data row 3+
    SB layout:  hint row 1, totals row 2, header row 3, data row 4+
    Columns 1..8: Symbol, Investment, Security, Qty, Open date,
                  Cost per share, Cost, Wash sale adj.
    """
    wb = openpyxl.Workbook()
    # Drop the auto-created blank sheet — we'll add named ones explicitly.
    default = wb.active
    if default is not None:
        wb.remove(default)

    if include_td:
        ws = wb.create_sheet("TD GainLoss Raw")
        # Row 1 stub.
        ws.cell(row=1, column=2, value=2000.0)
        # Row 2 header.
        for col, hdr in enumerate(
            ["Symbol", "Investment", "Security", "Qty", "Open date",
             "Adj cost per share", "Adj cost", "Wash sale adj"],
            start=1,
        ):
            ws.cell(row=2, column=col, value=hdr)
        # Row 3+ data.
        ws.cell(row=3, column=1, value="BIV")
        ws.cell(row=3, column=2, value=5997.8)
        ws.cell(row=3, column=3, value="VANGUARD INTERMEDIATE-TERM (BIV)")
        ws.cell(row=3, column=4, value=69.0)
        ws.cell(row=3, column=5, value=datetime(2015, 1, 29))
        ws.cell(row=3, column=6, value=86.9246)
        ws.cell(row=3, column=7, value=5997.8)
        ws.cell(row=3, column=8, value=None)

        ws.cell(row=4, column=1, value="BIV")
        ws.cell(row=4, column=2, value=0)
        ws.cell(row=4, column=3, value="VANGUARD INTERMEDIATE-TERM (BIV)")
        ws.cell(row=4, column=4, value=0.149)
        ws.cell(row=4, column=5, value=datetime(2015, 11, 6))
        ws.cell(row=4, column=6, value=83.2886)
        ws.cell(row=4, column=7, value=12.41)
        ws.cell(row=4, column=8, value=None)

        if include_na_rows:
            # '#N/A' symbol → skipped.
            ws.cell(row=5, column=1, value="#N/A")
            ws.cell(row=5, column=2, value=0)
            # Symbol present but qty is '#N/A' → skipped.
            ws.cell(row=6, column=1, value="VTI")
            ws.cell(row=6, column=3, value="VANGUARD TOTAL")
            ws.cell(row=6, column=4, value="#N/A")
            ws.cell(row=6, column=5, value=datetime(2018, 6, 1))
            ws.cell(row=6, column=6, value=140.0)
            ws.cell(row=6, column=7, value="#N/A")

    if include_sb:
        ws = wb.create_sheet("SB Raw")
        ws.cell(row=1, column=1, value="Where to Find: ...")
        ws.cell(row=2, column=2, value=37869.96)
        for col, hdr in enumerate(
            ["Symbol", "Investment", "Security", "Qty", "Open date",
             "Cost per share", "Cost", "Wash sale adj"],
            start=1,
        ):
            ws.cell(row=3, column=col, value=hdr)

        ws.cell(row=4, column=1, value="XLE")
        ws.cell(row=4, column=2, value=461.87)
        ws.cell(row=4, column=3, value="ENERGY SECTOR INDEX SPDR (XLE)")
        ws.cell(row=4, column=4, value=8.4108)
        ws.cell(row=4, column=5, value=datetime(2009, 4, 7))
        ws.cell(row=4, column=6, value=54.9139)
        ws.cell(row=4, column=7, value=461.87)
        ws.cell(row=4, column=8, value=86.87)

        ws.cell(row=5, column=1, value="XLE")
        ws.cell(row=5, column=2, value=0)
        ws.cell(row=5, column=3, value="ENERGY SECTOR INDEX SPDR (XLE)")
        ws.cell(row=5, column=4, value=0.0443)
        ws.cell(row=5, column=5, value=datetime(2009, 7, 1))
        ws.cell(row=5, column=6, value=48.0813)
        ws.cell(row=5, column=7, value=2.13)
        ws.cell(row=5, column=8, value=None)

        if include_na_rows:
            # Empty symbol row → skipped (real SB sheet has these as trailing).
            ws.cell(row=6, column=1, value=None)
            ws.cell(row=6, column=2, value=0)
            # Symbol with '#N/A' cost → skipped.
            ws.cell(row=7, column=1, value="MSFT")
            ws.cell(row=7, column=3, value="MICROSOFT")
            ws.cell(row=7, column=4, value=10.0)
            ws.cell(row=7, column=5, value=datetime(2020, 1, 2))
            ws.cell(row=7, column=6, value="#N/A")
            ws.cell(row=7, column=7, value="#N/A")

    wb.save(path)


@pytest.fixture
def lots_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "lots_fixture.xlsx"
    _build_lots_workbook(p)
    return p


@pytest.fixture
def lots_xlsx_td_only(tmp_path: Path) -> Path:
    p = tmp_path / "lots_td_only.xlsx"
    _build_lots_workbook(p, include_sb=False)
    return p


@pytest.fixture
def lots_xlsx_sb_only(tmp_path: Path) -> Path:
    p = tmp_path / "lots_sb_only.xlsx"
    _build_lots_workbook(p, include_td=False)
    return p


@pytest.fixture
def lots_xlsx_with_na(tmp_path: Path) -> Path:
    p = tmp_path / "lots_with_na.xlsx"
    _build_lots_workbook(p, include_na_rows=True)
    return p


def test_import_lots_td_sheet(session: Session, lots_xlsx_td_only: Path) -> None:
    """TD-only fixture: two real lot rows; raw_account_name='TD Ameritrade'."""
    result = import_cost_basis_lots(
        str(lots_xlsx_td_only), dry_run=False, session=session
    )

    assert result.imported == 2, f"unexpected count: {result}"
    rows = session.query(CostBasisLot).all()
    assert len(rows) == 2
    assert all(r.raw_account_name == TD_RAW_ACCOUNT_NAME for r in rows)
    assert all(r.source == TD_LOTS_SOURCE_TAG for r in rows)
    assert all(r.account_id is None for r in rows)
    assert {r.symbol for r in rows} == {"BIV"}
    # Quantity precision preserved as Decimal.
    quantities = sorted(r.quantity for r in rows)
    assert quantities == [Decimal("0.149"), Decimal("69")]


def test_import_lots_sb_sheet(session: Session, lots_xlsx_sb_only: Path) -> None:
    """SB-only fixture: two real lot rows; raw_account_name='Sharebuilder'."""
    result = import_cost_basis_lots(
        str(lots_xlsx_sb_only), dry_run=False, session=session
    )

    assert result.imported == 2, f"unexpected count: {result}"
    rows = session.query(CostBasisLot).all()
    assert len(rows) == 2
    assert all(r.raw_account_name == SB_RAW_ACCOUNT_NAME for r in rows)
    assert all(r.source == SB_LOTS_SOURCE_TAG for r in rows)
    assert {r.symbol for r in rows} == {"XLE"}
    # First XLE row carries a wash_sale_adj; second is None.
    wash = sorted(
        (r.wash_sale_adj for r in rows),
        key=lambda x: (x is None, x),
    )
    assert wash[0] == Decimal("86.87")
    assert wash[1] is None


def test_import_lots_dedup_by_row_hash(
    session: Session, lots_xlsx: Path
) -> None:
    """Re-running the same workbook produces 0 new rows."""
    first = import_cost_basis_lots(
        str(lots_xlsx), dry_run=False, session=session
    )
    # 2 TD + 2 SB = 4
    assert first.imported == 4

    second = import_cost_basis_lots(
        str(lots_xlsx), dry_run=False, session=session
    )
    assert second.imported == 0
    assert second.dup_skipped == 4
    assert session.query(CostBasisLot).count() == 4


def test_import_lots_skips_na(
    session: Session, lots_xlsx_with_na: Path
) -> None:
    """Rows with '#N/A' symbol/qty/cost are skipped silently."""
    result = import_cost_basis_lots(
        str(lots_xlsx_with_na), dry_run=False, session=session
    )

    # Same 2 TD + 2 SB real lots — N/A rows ignored.
    assert result.imported == 4, f"unexpected count: {result}"
    assert result.errors == []
    rows = session.query(CostBasisLot).all()
    symbols = {r.symbol for r in rows}
    assert symbols == {"BIV", "XLE"}
    # Confirmed: VTI (TD #N/A qty) and MSFT (SB #N/A cost) are absent.
    assert "VTI" not in symbols
    assert "MSFT" not in symbols
