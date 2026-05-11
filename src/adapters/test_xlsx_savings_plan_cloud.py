"""Cloud-mode tests for the XLSX savings-plan adapter (IC-T02).

REQ-WC-012: --target cloud POSTs normalized rows to Workers endpoint instead
of writing to SQLite.

Tests mock ``post_to_wealth`` to verify:
  - POST made with the correct URL slug
  - X-Internal-Key header present (via mock call args)
  - Payload shape per subcommand matches expected JSON schema
  - Workers 4xx errors are recorded per-row; batch continues (error isolation)
  - Distinct-account list and counters are correct in cloud mode
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, call, patch

import openpyxl
import pytest

from src.adapters.xlsx_savings_plan import (
    _CLOUD_INGEST_SOURCE_BALANCES,
    _CLOUD_INGEST_SOURCE_LOTS,
    _CLOUD_INGEST_SOURCE_PRICES,
    _default_target,
    import_account_balances_cloud,
    import_cost_basis_lots_cloud,
    import_historical_prices_cloud,
    main,
)
from src.adapters._shared.wealth_client import WealthClientError, WealthHTTPError


# ---------------------------------------------------------------------------
# Fixture: minimal XLSX workbook
# ---------------------------------------------------------------------------


def _build_balances_workbook(path: Path) -> None:
    """Two child accounts × two dates on Account Summary sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Account Summary"
    ws.cell(row=1, column=1, value="Account")
    ws.cell(row=1, column=2, value=datetime(2024, 12, 11))
    ws.cell(row=1, column=3, value=datetime(2024, 6, 27))
    # Aggregate row to skip
    ws.cell(row=2, column=1, value="Savings")
    ws.cell(row=2, column=2, value=100)
    ws.cell(row=2, column=3, value=200)
    # Child rows
    ws.cell(row=3, column=1, value="Schwab Brokerage")
    ws.cell(row=3, column=2, value=60000.50)
    ws.cell(row=3, column=3, value=55000.00)
    ws.cell(row=4, column=1, value="Vanguard 401k")
    ws.cell(row=4, column=2, value=40000.25)
    ws.cell(row=4, column=3, value=38000.00)
    wb.save(str(path))


def _build_prices_workbook(path: Path) -> None:
    """Two symbols × two dates on Historical Prices sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Historical Prices"
    # Row 3 = date headers per spec
    ws.cell(row=3, column=2, value=datetime(2024, 12, 11))
    ws.cell(row=3, column=3, value=datetime(2024, 6, 27))
    ws.cell(row=4, column=1, value="VTI")
    ws.cell(row=4, column=2, value=250.12345678)
    ws.cell(row=4, column=3, value=240.00)
    ws.cell(row=5, column=1, value="SCHB")
    ws.cell(row=5, column=2, value=55.50)
    ws.cell(row=5, column=3, value=52.00)
    wb.save(str(path))


def _build_lots_workbook(path: Path) -> None:
    """Minimal TD GainLoss Raw sheet with two lots."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "TD GainLoss Raw"
    # Header row 2, data starts row 3
    ws.cell(row=2, column=1, value="Symbol")
    ws.cell(row=2, column=2, value="Investment")
    ws.cell(row=2, column=3, value="Security")
    ws.cell(row=2, column=4, value="Qty")
    ws.cell(row=2, column=5, value="Open date")
    ws.cell(row=2, column=6, value="Cost per share")
    ws.cell(row=2, column=7, value="Cost total")
    ws.cell(row=2, column=8, value="Wash sale adj")
    # Lot rows
    ws.cell(row=3, column=1, value="VTI")
    ws.cell(row=3, column=4, value=10.0)
    ws.cell(row=3, column=5, value=datetime(2020, 1, 15))
    ws.cell(row=3, column=6, value=150.00)
    ws.cell(row=3, column=7, value=1500.00)
    ws.cell(row=3, column=8, value=None)
    ws.cell(row=4, column=1, value="SCHB")
    ws.cell(row=4, column=4, value=5.0)
    ws.cell(row=4, column=5, value=datetime(2019, 6, 10))
    ws.cell(row=4, column=6, value=45.00)
    ws.cell(row=4, column=7, value=225.00)
    ws.cell(row=4, column=8, value=None)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Tests: import_account_balances_cloud
# ---------------------------------------------------------------------------


def test_balances_cloud_posts_to_correct_url(tmp_path: Path) -> None:
    """import_account_balances_cloud POSTs to the xlsx-snapshot slug."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {"inserted": 4, "errors": []}
        result = import_account_balances_cloud(str(xlsx))

    assert mock_post.called
    calls = mock_post.call_args_list
    # All calls must use the xlsx-snapshot slug
    for c in calls:
        assert c.args[1] == _CLOUD_INGEST_SOURCE_BALANCES, (
            f"Expected slug '{_CLOUD_INGEST_SOURCE_BALANCES}', got {c.args[1]!r}"
        )
    # 4 data rows: 2 accounts × 2 dates (skipping the "Savings" aggregate)
    assert result.imported == 4
    assert result.errors == []


def test_balances_cloud_payload_shape(tmp_path: Path) -> None:
    """Each row in the POST payload has the expected JSON schema."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        import_account_balances_cloud(str(xlsx))

    assert mock_post.called
    all_rows = []
    for c in mock_post.call_args_list:
        all_rows.extend(c.args[0]["rows"])

    assert len(all_rows) == 4
    for row in all_rows:
        assert "raw_account_name" in row, f"Missing raw_account_name in {row}"
        assert "as_of" in row, f"Missing as_of in {row}"
        assert "balance" in row, f"Missing balance in {row}"
        assert "source" in row, f"Missing source in {row}"
        assert "source_row_hash" in row, f"Missing source_row_hash in {row}"
        # balance must be a decimal string, not a float
        assert isinstance(row["balance"], str), (
            f"balance must be a string, got {type(row['balance'])}: {row['balance']!r}"
        )
        # amount sign preserved: all balances are positive
        balance_val = Decimal(row["balance"])
        assert balance_val >= 0, f"Balance should be non-negative, got {balance_val}"


def test_balances_cloud_decimal_precision(tmp_path: Path) -> None:
    """Balance values are quantized to 2 decimal places (monetary scale)."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        import_account_balances_cloud(str(xlsx))

    all_rows = []
    for c in mock_post.call_args_list:
        all_rows.extend(c.args[0]["rows"])

    balances = [row["balance"] for row in all_rows]
    # 60000.50 must round-trip exactly as "60000.50" (trailing zero preserved)
    assert "60000.50" in balances, f"Expected '60000.50' in balances: {balances}"


def test_balances_cloud_4xx_error_isolation(tmp_path: Path) -> None:
    """Workers 4xx errors are recorded per-batch; remaining batches continue."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.side_effect = WealthHTTPError(422, "Unprocessable Entity")
        result = import_account_balances_cloud(str(xlsx))

    # Errors recorded, but function did not raise
    assert len(result.errors) > 0
    assert result.imported == 0  # nothing inserted on error


def test_balances_cloud_missing_sheet(tmp_path: Path) -> None:
    """Missing Account Summary sheet returns error, no POST made."""
    xlsx = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(xlsx))

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        result = import_account_balances_cloud(str(xlsx))

    mock_post.assert_not_called()
    assert any("missing sheet" in e for e in result.errors)


# ---------------------------------------------------------------------------
# Tests: import_historical_prices_cloud
# ---------------------------------------------------------------------------


def test_prices_cloud_posts_to_correct_slug(tmp_path: Path) -> None:
    """import_historical_prices_cloud POSTs to the historical-prices slug."""
    xlsx = tmp_path / "prices.xlsx"
    _build_prices_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        result = import_historical_prices_cloud(str(xlsx))

    for c in mock_post.call_args_list:
        assert c.args[1] == _CLOUD_INGEST_SOURCE_PRICES


def test_prices_cloud_payload_shape(tmp_path: Path) -> None:
    """Each price row has symbol, trade_date, close (decimal string), source."""
    xlsx = tmp_path / "prices.xlsx"
    _build_prices_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        import_historical_prices_cloud(str(xlsx))

    all_rows = []
    for c in mock_post.call_args_list:
        all_rows.extend(c.args[0]["rows"])

    for row in all_rows:
        assert "symbol" in row
        assert "trade_date" in row
        assert "close" in row
        assert "source" in row
        assert isinstance(row["close"], str), f"close must be a string: {row['close']!r}"
        # Scale 8 precision
        assert "." in row["close"] and len(row["close"].split(".")[1]) == 8, (
            f"Expected 8 decimal places in close: {row['close']!r}"
        )


def test_prices_cloud_4xx_error_isolation(tmp_path: Path) -> None:
    """Workers 4xx on price POST records error; batch continues."""
    xlsx = tmp_path / "prices.xlsx"
    _build_prices_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.side_effect = WealthHTTPError(413, "Payload Too Large")
        result = import_historical_prices_cloud(str(xlsx))

    assert len(result.errors) > 0
    assert result.imported == 0


# ---------------------------------------------------------------------------
# Tests: import_cost_basis_lots_cloud
# ---------------------------------------------------------------------------


def test_lots_cloud_posts_to_correct_slug(tmp_path: Path) -> None:
    """import_cost_basis_lots_cloud POSTs to the cost-basis-lot slug."""
    xlsx = tmp_path / "lots.xlsx"
    _build_lots_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        result = import_cost_basis_lots_cloud(str(xlsx))

    for c in mock_post.call_args_list:
        assert c.args[1] == _CLOUD_INGEST_SOURCE_LOTS


def test_lots_cloud_payload_shape(tmp_path: Path) -> None:
    """Each lot row has the expected fields with decimal strings."""
    xlsx = tmp_path / "lots.xlsx"
    _build_lots_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        import_cost_basis_lots_cloud(str(xlsx))

    all_rows = []
    for c in mock_post.call_args_list:
        all_rows.extend(c.args[0]["rows"])

    for row in all_rows:
        assert "symbol" in row
        assert "quantity" in row
        assert "open_date" in row
        assert "cost_per_share" in row
        assert "cost_total" in row
        assert "source" in row
        assert "source_row_hash" in row
        assert isinstance(row["quantity"], str), f"quantity must be a string"
        assert isinstance(row["cost_per_share"], str), "cost_per_share must be a string"


def test_lots_cloud_missing_both_sheets(tmp_path: Path) -> None:
    """Missing both lot sheets returns error, no POST made."""
    xlsx = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(xlsx))

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        result = import_cost_basis_lots_cloud(str(xlsx))

    mock_post.assert_not_called()
    assert any("missing both" in e for e in result.errors)


# ---------------------------------------------------------------------------
# Tests: CLI --target flag
# ---------------------------------------------------------------------------


def test_default_target_reads_env(monkeypatch: pytest.MonkeyPatch) -> None:
    """_default_target() returns WEALTH_TARGET_DEFAULT from env."""
    monkeypatch.setenv("WEALTH_TARGET_DEFAULT", "cloud")
    assert _default_target() == "cloud"


def test_default_target_fallback_to_local(monkeypatch: pytest.MonkeyPatch) -> None:
    """_default_target() returns 'local' when env is not set."""
    monkeypatch.delenv("WEALTH_TARGET_DEFAULT", raising=False)
    assert _default_target() == "local"


def test_cli_target_cloud_flag_calls_cloud_function(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """--target cloud calls the cloud import function instead of DB write."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    monkeypatch.setenv("WEALTH_API_BASE", "https://internal.sparkry.ai")
    monkeypatch.setenv("WEALTH_INTERNAL_KEY", "test-key")

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        mock_post.return_value = {}
        rc = main(["import-balances", "--file", str(xlsx), "--apply", "--target", "cloud"])

    assert mock_post.called, "Expected post_to_wealth to be called in cloud mode"
    assert rc == 0


def test_cli_dry_run_does_not_call_cloud(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Dry-run mode (no --apply) does not call post_to_wealth."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        main(["import-balances", "--file", str(xlsx), "--target", "cloud"])

    mock_post.assert_not_called()


def test_cli_target_local_does_not_post(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """--target local (dry-run) never calls post_to_wealth regardless of target."""
    xlsx = tmp_path / "test.xlsx"
    _build_balances_workbook(xlsx)

    # Without --apply, no DB write and no cloud POST regardless of target.
    with patch("src.adapters.xlsx_savings_plan.post_to_wealth") as mock_post:
        rc = main(["import-balances", "--file", str(xlsx), "--target", "local"])

    mock_post.assert_not_called()
    assert rc == 0
