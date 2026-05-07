"""XLSX importer for the user's "Savings & Retirement Plan" workbook.

Reads the "Account Summary" sheet — a wide-form pivot of account balances over
time. Header row column 0 is the literal "Account"; columns 1..N are
``datetime`` objects (snapshot dates). Body rows have an account label in
column 0 and per-date balances in columns 1..N (or None).

Top-level aggregate rows ("Savings", "Retirement", "Total", "College Savings")
are sums of indented child rows and are SKIPPED to avoid double-counting.
Projection scratch rows ("A Projected", "E Projected", "Rate", "Yearly",
"Years", "Annual") are also skipped.

Each (account, date, balance) tuple is written to ``account_balance_snapshot``
with ``source = 'xlsx_2024'``. The importer is idempotent: re-runs over the
same workbook produce zero new rows because the row-hash and the natural-key
UNIQUE constraints both reject duplicates.

Account-name → Account mapping is intentionally NOT done here. A separate
review step assigns ``account_id`` to ``raw_account_name`` values; for now
every row lands with ``account_id = NULL``.
"""

from __future__ import annotations

import argparse
import contextlib
import hashlib
import logging
import sys
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING

import openpyxl
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from src.models.enums import IngestionStatus
from src.models.history import AccountBalanceSnapshot, CostBasisLot, HistoricalPrice
from src.models.ingestion_log import IngestionLog

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ── Constants ────────────────────────────────────────────────────────────────

SOURCE_TAG = "xlsx_2024"
"""Value written to ``account_balance_snapshot.source``."""

ADAPTER_NAME = "xlsx_savings_plan"
"""Identifier written to ``ingestion_log.source``."""

SHEET_NAME = "Account Summary"

PRICES_SHEET_NAME = "Historical Prices"
"""Sheet that seeds ``historical_price`` rows."""

PRICES_SOURCE_TAG = "xlsx_2024_prices"
"""Distinct from ``SOURCE_TAG`` so the source column lets us filter "XLSX
balances" vs "XLSX prices" independently for downstream cleanup."""

TD_LOTS_SHEET_NAME = "TD GainLoss Raw"
SB_LOTS_SHEET_NAME = "SB Raw"

TD_LOTS_SOURCE_TAG = "xlsx_td_gainloss"
SB_LOTS_SOURCE_TAG = "xlsx_sb_raw"

TD_RAW_ACCOUNT_NAME = "TD Ameritrade"
SB_RAW_ACCOUNT_NAME = "Sharebuilder"

# In TD GainLoss Raw, header is row 2 and data starts at row 3.
# In SB Raw, row 1 is a "Where to Find" hint, row 2 is a totals row, header
# is row 3, and data starts at row 4.
_TD_HEADER_ROW = 2
_SB_HEADER_ROW = 3

# Names (case-insensitive, stripped) that are aggregate / projection rows
# we don't want as snapshots. Aggregates would double-count their children.
SKIP_NAMES: frozenset[str] = frozenset(
    n.lower()
    for n in (
        "Savings",
        "Retirement",
        "Total",
        "College Savings",
        "A Projected",
        "E Projected",
        "Rate",
        "Yearly",
        "Years",
        "Annual",
    )
)


# ── Result dataclass ─────────────────────────────────────────────────────────


@dataclass
class ImportResult:
    """Summary of an import run."""

    imported: int = 0
    """Newly inserted snapshot rows."""

    matched: int = 0
    """Rows whose account_id was resolved to a live Account (currently 0)."""

    unmatched: int = 0
    """Rows inserted with account_id = NULL."""

    dup_skipped: int = 0
    """Rows skipped because an equivalent snapshot already exists."""

    errors: list[str] = field(default_factory=list)
    """Per-row error strings (record_label: message)."""

    distinct_accounts: list[str] = field(default_factory=list)
    """Distinct ``raw_account_name`` values observed (excluding skipped)."""


# ── Helpers ──────────────────────────────────────────────────────────────────

# Quantize Decimals before hashing: same economic value with different
# trailing-zero representations (Decimal('10.50') vs '10.5') would otherwise
# hash differently and break re-import idempotence.
_QTY_QUANT = Decimal("0.00000001")
_MONEY_QUANT = Decimal("0.01")


def _row_hash(raw_account_name: str, as_of: date, balance: Decimal) -> str:
    """SHA256 hex of the canonical row identity tuple."""
    payload = (
        f"{raw_account_name}|{as_of.isoformat()}"
        f"|{balance.quantize(_MONEY_QUANT)}|{SOURCE_TAG}"
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _coerce_date(value: object) -> date | None:
    """Coerce a header cell value to a ``date``."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _coerce_balance(value: object) -> Decimal | None:
    """Coerce a body cell value to ``Decimal``.

    None, blank strings, and non-numeric values (e.g. ``'#N/A'``) return None.
    Floats and ints are stringified first to preserve the user's intended
    decimal representation rather than dragging in float artefacts.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        # bool is an int subclass — exclude.
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # str(float) gives the shortest round-trippable repr — safer than
        # Decimal(float) which would expose binary-float noise.
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "")
        if not cleaned or cleaned in {"#N/A", "-", "--", "n/a", "N/A"}:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    return None


def _is_skip_row(name: str) -> bool:
    """True if ``name`` is an aggregate/projection row we don't snapshot."""
    return name.strip().lower() in SKIP_NAMES


# ── Core import ──────────────────────────────────────────────────────────────


def _read_header_dates(ws: Worksheet) -> list[tuple[int, date]]:
    """Return ``[(col_index, snapshot_date), ...]`` for date columns in the header."""
    out: list[tuple[int, date]] = []
    for col_idx in range(2, ws.max_column + 1):  # 1-based; col 1 is "Account"
        cell = ws.cell(row=1, column=col_idx)
        d = _coerce_date(cell.value)
        if d is not None:
            out.append((col_idx, d))
    return out


def _iter_snapshot_rows(
    ws: Worksheet, date_cols: list[tuple[int, date]]
) -> list[tuple[str, date, Decimal]]:
    """Yield (raw_account_name, as_of, balance) tuples for every valid cell.

    Skips aggregate rows by name; skips cells whose value isn't numeric.
    """
    out: list[tuple[str, date, Decimal]] = []
    for row_idx in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row_idx, column=1).value
        if not isinstance(name_cell, str):
            continue
        raw_account_name = name_cell.strip()
        if not raw_account_name:
            continue
        if _is_skip_row(raw_account_name):
            continue
        for col_idx, as_of in date_cols:
            bal = _coerce_balance(ws.cell(row=row_idx, column=col_idx).value)
            if bal is None:
                continue
            out.append((raw_account_name, as_of, bal))
    return out


def import_account_balances(
    file_path: str,
    dry_run: bool = True,
    session: Session | None = None,
) -> ImportResult:
    """Import the Account Summary sheet of the savings/retirement workbook.

    Args:
        file_path: Path to the XLSX workbook.
        dry_run:   When True, parse-and-count only; never write to ``session``.
                   Default True to protect the live DB during exploration.
        session:   SQLAlchemy session. Required when ``dry_run`` is False.

    Returns:
        :class:`ImportResult` with counts and distinct account names.
    """
    result = ImportResult()

    # We don't use read_only=True: it leaves max_row/max_column lazy (None until
    # the sheet is iterated), which complicates indexed access by (row, col).
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        result.errors.append(f"workbook missing sheet '{SHEET_NAME}'")
        if session is not None and not dry_run:
            _log_run(session, result, status=IngestionStatus.FAILURE,
                     error_detail=result.errors[-1])
        return result

    ws = wb[SHEET_NAME]
    date_cols = _read_header_dates(ws)
    if not date_cols:
        result.errors.append("no date columns found in header row")
        if session is not None and not dry_run:
            _log_run(session, result, status=IngestionStatus.FAILURE,
                     error_detail=result.errors[-1])
        return result

    rows = _iter_snapshot_rows(ws, date_cols)

    distinct = sorted({name for name, _, _ in rows})
    result.distinct_accounts = distinct

    if dry_run:
        # Count would-be-inserted rows; we can't know dup_skipped without DB.
        result.imported = 0
        result.unmatched = len(rows)
        return result

    if session is None:
        result.errors.append("session required when dry_run=False")
        return result

    # ── Apply ────────────────────────────────────────────────────────────────
    # Per-row savepoint isolates IntegrityError without losing earlier rows;
    # outer commit batches the whole sheet into one fsync at the end.
    for raw_account_name, as_of, balance in rows:
        record_label = f"{raw_account_name}@{as_of.isoformat()}"
        row_hash = _row_hash(raw_account_name, as_of, balance)

        existing = (
            session.query(AccountBalanceSnapshot.id)
            .filter(AccountBalanceSnapshot.source_row_hash == row_hash)
            .first()
        )
        if existing is not None:
            result.dup_skipped += 1
            continue

        try:
            with session.begin_nested():
                session.add(
                    AccountBalanceSnapshot(
                        account_id=None,
                        raw_account_name=raw_account_name,
                        as_of=as_of,
                        balance=balance,
                        source=SOURCE_TAG,
                        source_row_hash=row_hash,
                    )
                )
            result.imported += 1
            result.unmatched += 1  # account_id is always NULL in this pass
        except IntegrityError:
            # UNIQUE on natural key — same logical row already exists.
            result.dup_skipped += 1
        except Exception as exc:  # noqa: BLE001 — per-record isolation
            result.errors.append(f"{record_label}: {exc}")
            logger.warning("xlsx_savings_plan: row %s failed: %s",
                           record_label, exc, exc_info=True)

    session.commit()

    # ── Audit log ────────────────────────────────────────────────────────────
    status = (
        IngestionStatus.SUCCESS
        if not result.errors
        else IngestionStatus.PARTIAL_FAILURE
    )
    _log_run(session, result, status=status,
             error_detail="\n".join(result.errors) or None)

    return result


def _log_run(
    session: Session,
    result: ImportResult,
    *,
    status: IngestionStatus,
    error_detail: str | None,
) -> None:
    """Record an IngestionLog entry. Failures here are swallowed (log-only)."""
    try:
        log = IngestionLog(
            source=ADAPTER_NAME,
            status=status.value,
            records_processed=result.imported + result.dup_skipped,
            records_failed=len(result.errors),
            error_detail=error_detail,
        )
        session.add(log)
        session.commit()
    except Exception:  # noqa: BLE001
        logger.exception("failed to write IngestionLog for %s", ADAPTER_NAME)
        with contextlib.suppress(Exception):
            session.rollback()


# ── Historical Prices ────────────────────────────────────────────────────────


def _price_row_hash(symbol: str, trade_date: date, close: Decimal) -> str:
    """SHA256 hex of the canonical price-row identity tuple."""
    payload = (
        f"{symbol}|{trade_date.isoformat()}"
        f"|{close.quantize(_QTY_QUANT)}|{PRICES_SOURCE_TAG}"
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _read_price_date_columns(ws: Worksheet) -> list[tuple[int, date]]:
    """Return ``[(col_index, trade_date), ...]`` from row 3 of Historical Prices.

    Row 3 holds the *weekend-adjusted* dates (per the workbook's own labelling
    in row 2 column A). Row 1 has bucket labels ("Yesterday", "30 Days", ...)
    that are ignored.
    """
    out: list[tuple[int, date]] = []
    for col_idx in range(2, ws.max_column + 1):
        d = _coerce_date(ws.cell(row=3, column=col_idx).value)
        if d is not None:
            out.append((col_idx, d))
    return out


def _iter_price_rows(
    ws: Worksheet, date_cols: list[tuple[int, date]]
) -> list[tuple[str, date, Decimal]]:
    """Yield (symbol, trade_date, close) tuples for every numeric price cell.

    Skips header rows (1-3), blank symbol rows, '#N/A' / non-numeric cells.
    """
    out: list[tuple[str, date, Decimal]] = []
    for row_idx in range(4, ws.max_row + 1):
        sym_cell = ws.cell(row=row_idx, column=1).value
        if not isinstance(sym_cell, str):
            continue
        symbol = sym_cell.strip().upper()
        if not symbol or symbol == "#N/A":
            continue
        for col_idx, trade_date in date_cols:
            close = _coerce_balance(ws.cell(row=row_idx, column=col_idx).value)
            if close is None:
                continue
            out.append((symbol, trade_date, close))
    return out


def import_historical_prices(
    file_path: str,
    dry_run: bool = True,
    session: Session | None = None,
) -> ImportResult:
    """Import the 'Historical Prices' sheet into ``historical_price``.

    Per row, emits up to four (symbol, trade_date, close) candidates — one per
    weekend-adjusted date column from row 3. Cells that are ``#N/A``, ``None``,
    or non-numeric are skipped silently.

    yfinance is the authoritative source for prices; this XLSX seed only fills
    historical gaps. ``IntegrityError`` (composite-PK collision on
    ``(symbol, trade_date)``) is therefore counted as ``dup_skipped`` rather
    than overwritten.

    Args mirror :func:`import_account_balances`.
    """
    result = ImportResult()

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if PRICES_SHEET_NAME not in wb.sheetnames:
        result.errors.append(f"workbook missing sheet '{PRICES_SHEET_NAME}'")
        if session is not None and not dry_run:
            _log_run(session, result, status=IngestionStatus.FAILURE,
                     error_detail=result.errors[-1])
        return result

    ws = wb[PRICES_SHEET_NAME]
    date_cols = _read_price_date_columns(ws)
    if not date_cols:
        result.errors.append("no date columns found in Historical Prices row 3")
        if session is not None and not dry_run:
            _log_run(session, result, status=IngestionStatus.FAILURE,
                     error_detail=result.errors[-1])
        return result

    rows = _iter_price_rows(ws, date_cols)

    distinct = sorted({sym for sym, _, _ in rows})
    result.distinct_accounts = distinct  # reusing field for "distinct symbols"

    if dry_run:
        result.imported = 0
        result.unmatched = len(rows)
        return result

    if session is None:
        result.errors.append("session required when dry_run=False")
        return result

    for symbol, trade_date, close in rows:
        record_label = f"{symbol}@{trade_date.isoformat()}"
        existing = (
            session.query(HistoricalPrice.symbol)
            .filter(
                HistoricalPrice.symbol == symbol,
                HistoricalPrice.trade_date == trade_date,
            )
            .first()
        )
        if existing is not None:
            result.dup_skipped += 1
            continue
        try:
            with session.begin_nested():
                session.add(
                    HistoricalPrice(
                        symbol=symbol,
                        trade_date=trade_date,
                        close=close,
                        source=PRICES_SOURCE_TAG,
                    )
                )
            result.imported += 1
        except IntegrityError:
            result.dup_skipped += 1
        except Exception as exc:  # noqa: BLE001 — per-record isolation
            result.errors.append(f"{record_label}: {exc}")
            logger.warning("xlsx_savings_plan.prices: row %s failed: %s",
                           record_label, exc, exc_info=True)

    session.commit()

    status = (
        IngestionStatus.SUCCESS
        if not result.errors
        else IngestionStatus.PARTIAL_FAILURE
    )
    _log_run(session, result, status=status,
             error_detail="\n".join(result.errors) or None)

    return result


# ── Cost Basis Lots ──────────────────────────────────────────────────────────


@dataclass
class _LotRow:
    """A normalised lot row before insertion."""

    symbol: str
    security_name: str | None
    quantity: Decimal
    open_date: date
    cost_per_share: Decimal
    cost_total: Decimal
    wash_sale_adj: Decimal | None
    row_idx: int


def _lot_row_hash(
    symbol: str,
    open_date: date,
    quantity: Decimal,
    cost_total: Decimal,
    source: str,
    row_idx: int,
) -> str:
    """SHA256 hex of the canonical lot-row identity tuple.

    ``row_idx`` is included so two lot rows with identical
    (symbol, open_date, quantity, cost_total) — common for RSU same-day
    tranches or repeated ETF buys at the same price — don't collide.
    """
    payload = (
        f"{symbol}|{open_date.isoformat()}"
        f"|{quantity.quantize(_QTY_QUANT)}"
        f"|{cost_total.quantize(_MONEY_QUANT)}"
        f"|{row_idx}|{source}"
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _parse_lot_row(
    ws: Worksheet,
    row_idx: int,
) -> _LotRow | None:
    """Parse a single lot-row from a TD/SB worksheet.

    The TD and SB sheets share column positions 1..8 for the fields we need:
      1=Symbol  2=Investment  3=Security  4=Qty  5=Open date
      6=Cost per share  7=Cost total  8=Wash sale adj

    Returns None for rows that are blank, header repeats, or have any
    required field missing/non-numeric/#N/A.
    """
    sym_cell = ws.cell(row=row_idx, column=1).value
    if not isinstance(sym_cell, str):
        return None
    symbol = sym_cell.strip().upper()
    if not symbol or symbol == "#N/A" or symbol == "SYMBOL":
        return None

    sec_cell = ws.cell(row=row_idx, column=3).value
    security_name = sec_cell.strip() if isinstance(sec_cell, str) else None

    qty = _coerce_balance(ws.cell(row=row_idx, column=4).value)
    od_raw = ws.cell(row=row_idx, column=5).value
    cps = _coerce_balance(ws.cell(row=row_idx, column=6).value)
    cost_total = _coerce_balance(ws.cell(row=row_idx, column=7).value)
    wash_adj = _coerce_balance(ws.cell(row=row_idx, column=8).value)

    open_date = _coerce_date(od_raw)

    if qty is None or cps is None or cost_total is None or open_date is None:
        return None

    return _LotRow(
        symbol=symbol,
        security_name=security_name,
        quantity=qty,
        open_date=open_date,
        cost_per_share=cps,
        cost_total=cost_total,
        wash_sale_adj=wash_adj,
        row_idx=row_idx,
    )


def _iter_lot_rows(ws: Worksheet, start_row: int) -> list[_LotRow]:
    """Walk a TD/SB worksheet from ``start_row`` collecting valid lot rows."""
    out: list[_LotRow] = []
    for row_idx in range(start_row, ws.max_row + 1):
        parsed = _parse_lot_row(ws, row_idx)
        if parsed is not None:
            out.append(parsed)
    return out


def _import_lots_from_sheet(
    ws: Worksheet,
    *,
    start_row: int,
    raw_account_name: str,
    source_tag: str,
    result: ImportResult,
    session: Session | None,
    dry_run: bool,
) -> None:
    """Shared core for TD and SB lot import. Mutates ``result`` in place."""
    rows = _iter_lot_rows(ws, start_row=start_row)

    # Track distinct symbols across both sheets via distinct_accounts.
    seen = set(result.distinct_accounts)
    for r in rows:
        seen.add(r.symbol)
    result.distinct_accounts = sorted(seen)

    if dry_run:
        result.unmatched += len(rows)
        return

    if session is None:
        result.errors.append("session required when dry_run=False")
        return

    for r in rows:
        record_label = f"{raw_account_name}:{r.symbol}@{r.open_date.isoformat()}"
        row_hash = _lot_row_hash(
            r.symbol, r.open_date, r.quantity, r.cost_total, source_tag, r.row_idx
        )

        existing = (
            session.query(CostBasisLot.id)
            .filter(CostBasisLot.source_row_hash == row_hash)
            .first()
        )
        if existing is not None:
            result.dup_skipped += 1
            continue

        try:
            with session.begin_nested():
                session.add(
                    CostBasisLot(
                        account_id=None,
                        raw_account_name=raw_account_name,
                        symbol=r.symbol,
                        security_name=r.security_name,
                        open_date=r.open_date,
                        quantity=r.quantity,
                        cost_per_share=r.cost_per_share,
                        cost_total=r.cost_total,
                        wash_sale_adj=r.wash_sale_adj,
                        source=source_tag,
                        source_row_hash=row_hash,
                    )
                )
            result.imported += 1
            result.unmatched += 1
        except IntegrityError:
            result.dup_skipped += 1
        except Exception as exc:  # noqa: BLE001 — per-record isolation
            result.errors.append(f"{record_label}: {exc}")
            logger.warning("xlsx_savings_plan.lots: row %s failed: %s",
                           record_label, exc, exc_info=True)
    # Outer commit happens in the caller after both TD and SB sheets process.


def import_cost_basis_lots(
    file_path: str,
    dry_run: bool = True,
    session: Session | None = None,
) -> ImportResult:
    """Import 'TD GainLoss Raw' and 'SB Raw' sheets into ``cost_basis_lot``.

    Both sheets follow the same first-eight-column layout (Symbol, Investment,
    Security, Qty, Open date, Cost per share, Cost, Wash sale adj). They differ
    only in start-row (TD's header is row 2, SB's is row 3) and provenance:

    - TD rows → ``raw_account_name='TD Ameritrade'``, ``source='xlsx_td_gainloss'``
    - SB rows → ``raw_account_name='Sharebuilder'``, ``source='xlsx_sb_raw'``

    Both are ingested in the same call; ``result`` aggregates counts. Account
    resolution (FK to ``account.id``) is deferred — ``account_id`` is left
    NULL and ``raw_account_name`` carries the audit trail.

    Idempotency is enforced via UNIQUE on ``cost_basis_lot.source_row_hash``,
    which is SHA256 of ``"{symbol}|{open_date}|{qty}|{cost_total}|{source}"``.
    """
    result = ImportResult()

    wb = openpyxl.load_workbook(file_path, data_only=True)

    found_any = False
    if TD_LOTS_SHEET_NAME in wb.sheetnames:
        found_any = True
        _import_lots_from_sheet(
            wb[TD_LOTS_SHEET_NAME],
            start_row=_TD_HEADER_ROW + 1,
            raw_account_name=TD_RAW_ACCOUNT_NAME,
            source_tag=TD_LOTS_SOURCE_TAG,
            result=result,
            session=session,
            dry_run=dry_run,
        )

    if SB_LOTS_SHEET_NAME in wb.sheetnames:
        found_any = True
        _import_lots_from_sheet(
            wb[SB_LOTS_SHEET_NAME],
            start_row=_SB_HEADER_ROW + 1,
            raw_account_name=SB_RAW_ACCOUNT_NAME,
            source_tag=SB_LOTS_SOURCE_TAG,
            result=result,
            session=session,
            dry_run=dry_run,
        )

    if not found_any:
        result.errors.append(
            f"workbook missing both '{TD_LOTS_SHEET_NAME}' and "
            f"'{SB_LOTS_SHEET_NAME}'"
        )
        if session is not None and not dry_run:
            _log_run(session, result, status=IngestionStatus.FAILURE,
                     error_detail=result.errors[-1])
        return result

    if dry_run:
        return result

    if session is None:
        # _import_lots_from_sheet would have already recorded the error.
        return result

    # Commit the batch of savepoint-isolated rows from both sheets.
    session.commit()

    status = (
        IngestionStatus.SUCCESS
        if not result.errors
        else IngestionStatus.PARTIAL_FAILURE
    )
    _log_run(session, result, status=status,
             error_detail="\n".join(result.errors) or None)

    return result


# ── CLI ──────────────────────────────────────────────────────────────────────


def _build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="python -m src.adapters.xlsx_savings_plan",
        description="Import sheets of the Savings & Retirement Plan XLSX.",
    )
    sub = p.add_subparsers(dest="cmd", required=True)

    for cmd_name, cmd_help in (
        ("import-balances",
         "Read the workbook's Account Summary sheet and import balance snapshots."),
        ("import-prices",
         "Read the workbook's Historical Prices sheet and seed historical_price."),
        ("import-lots",
         "Read 'TD GainLoss Raw' and 'SB Raw' and import cost-basis lots."),
    ):
        s = sub.add_parser(cmd_name, help=cmd_help)
        s.add_argument("--file", required=True, help="Path to the XLSX workbook.")
        s.add_argument(
            "--apply",
            action="store_true",
            help="Actually write to the live DB. Default is dry-run.",
        )
    return p


def _print_summary(result: ImportResult, dry_run: bool, *, label: str = "xlsx_savings_plan") -> None:
    mode = "DRY RUN" if dry_run else "APPLIED"
    print(f"=== {label} ({mode}) ===")
    print(f"  imported     : {result.imported}")
    print(f"  matched      : {result.matched}")
    print(f"  unmatched    : {result.unmatched}")
    print(f"  dup_skipped  : {result.dup_skipped}")
    print(f"  errors       : {len(result.errors)}")
    print(f"  distinct items ({len(result.distinct_accounts)}):")
    for name in result.distinct_accounts:
        print(f"    - {name}")
    if result.errors:
        print("  --- error detail ---")
        for e in result.errors[:20]:
            print(f"    * {e}")
        if len(result.errors) > 20:
            print(f"    ... {len(result.errors) - 20} more")


_CMD_DISPATCH = {
    "import-balances": (import_account_balances, "xlsx_savings_plan:balances"),
    "import-prices": (import_historical_prices, "xlsx_savings_plan:prices"),
    "import-lots": (import_cost_basis_lots, "xlsx_savings_plan:lots"),
}


def main(argv: list[str] | None = None) -> int:
    args = _build_arg_parser().parse_args(argv)
    if args.cmd not in _CMD_DISPATCH:
        return 2

    func, label = _CMD_DISPATCH[args.cmd]
    dry_run = not args.apply

    if dry_run:
        result = func(args.file, dry_run=True, session=None)
        _print_summary(result, dry_run=True, label=label)
        return 0

    # Live apply — open a session against the configured DB.
    try:
        from src.db.connection import get_session  # late import to keep tests light
    except ImportError as exc:  # pragma: no cover — environmental
        print(f"cannot import DB session factory: {exc}", file=sys.stderr)
        return 1

    with get_session() as session:
        result = func(args.file, dry_run=False, session=session)

    _print_summary(result, dry_run=False, label=label)
    return 0


if __name__ == "__main__":  # pragma: no cover — exercised via CLI
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise SystemExit(1) from None


__all__ = [
    "ADAPTER_NAME",
    "ImportResult",
    "PRICES_SOURCE_TAG",
    "SB_LOTS_SOURCE_TAG",
    "SB_RAW_ACCOUNT_NAME",
    "SOURCE_TAG",
    "TD_LOTS_SOURCE_TAG",
    "TD_RAW_ACCOUNT_NAME",
    "import_account_balances",
    "import_cost_basis_lots",
    "import_historical_prices",
]
